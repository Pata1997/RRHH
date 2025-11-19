from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file, current_app
from flask_login import login_required, current_user
from sqlalchemy import func, desc, or_
from decimal import Decimal
from datetime import datetime, date, timedelta
import calendar
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import json
import os
from werkzeug.utils import secure_filename
from io import BytesIO

from ..models import (
    db, Empleado, Cargo, Asistencia, Permiso, Sancion,
    Contrato, Liquidacion, Vacacion, IngresoExtra, Descuento,
    Bitacora, EstadoEmpleadoEnum, EstadoVacacionEnum, EstadoPermisoEnum, RoleEnum, Despido,
    Postulante, DocumentosCurriculum, AsistenciaEvento, Empresa, HorasExtra, Anticipo,
    SalarioMinimo, BonificacionFamiliar, TipoHijoEnum
)
from ..bitacora import registrar_bitacora, registrar_operacion_crud
from ..reports.report_utils import ReportUtils
from openpyxl import Workbook
from io import BytesIO as IOBytes

rrhh_bp = Blueprint('rrhh', __name__, url_prefix='/rrhh')

from functools import wraps

# Decorador para verificar rol
def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if current_user.rol not in roles:
                flash('No tienes permiso para acceder a este módulo', 'danger')
                return redirect(url_for('main.dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Configuraciones de uploads
UPLOADS_ROOT = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')
PERMISOS_UPLOAD_FOLDER = os.path.join(UPLOADS_ROOT, 'permisos')
SANCIONES_UPLOAD_FOLDER = os.path.join(UPLOADS_ROOT, 'sanciones')
POSTULANTES_UPLOAD_FOLDER = os.path.join(UPLOADS_ROOT, 'postulantes')
ANTICIPOS_UPLOAD_FOLDER = os.path.join(UPLOADS_ROOT, 'anticipos')
INGRESOS_UPLOAD_FOLDER = os.path.join(UPLOADS_ROOT, 'ingresos_extras')

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Servir archivos subidos (permisos/sanciones/postulantes)
@rrhh_bp.route('/uploads/<path:subpath>')
@login_required
def serve_uploads(subpath):
    # base debe apuntar a app/ para buscar app/uploads/...
    base = os.path.dirname(os.path.dirname(__file__))  # app/
    # Pero subpath viene como "uploads/permisos/archivo.pdf"
    # Queremos buscar en app/uploads/permisos/archivo.pdf
    # Así que no necesitamos el prefijo "uploads/" en subpath
    filepath = os.path.join(base, 'uploads', subpath)
    print(f"\n=== DEBUG serve_uploads ===")
    print(f"subpath: {subpath}")
    print(f"base: {base}")
    print(f"filepath: {filepath}")
    print(f"exists: {os.path.exists(filepath)}")
    if os.path.exists(filepath):
        print(f"✓ Sirviendo archivo: {filepath}")
    if not os.path.exists(filepath):
        print(f"ERROR: Archivo no encontrado")
        return jsonify({'error': 'File not found'}), 404
    return send_file(filepath)

# ==================== EMPLEADOS ====================
@rrhh_bp.route('/empleados', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def listar_empleados():
    """Lista todos los empleados"""
    registrar_bitacora(current_user, 'empleados', 'VIEW', 'empleados')
    
    pagina = request.args.get('page', 1, type=int)
    empleados = Empleado.query.paginate(page=pagina, per_page=10)
    
    return render_template('rrhh/empleados.html', empleados=empleados)

@rrhh_bp.route('/empleados/crear', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def crear_empleado():
    """Crear nuevo empleado"""
    cargos = Cargo.query.all()
    
    if request.method == 'POST':
        try:
            # Validar datos
            codigo = request.form.get('codigo')
            nombre = request.form.get('nombre')
            apellido = request.form.get('apellido')
            ci = request.form.get('ci')
            cargo_id = request.form.get('cargo_id')
            salario_base = request.form.get('salario_base')
            fecha_ingreso = datetime.strptime(request.form.get('fecha_ingreso'), '%Y-%m-%d').date()
            
            # Verificar que no exista
            if Empleado.query.filter_by(codigo=codigo).first():
                flash('El código del empleado ya existe', 'danger')
                return redirect(url_for('rrhh.crear_empleado'))
            
            if Empleado.query.filter_by(ci=ci).first():
                flash('El CI ya existe en el sistema', 'danger')
                return redirect(url_for('rrhh.crear_empleado'))
            
            # Crear empleado
            empleado = Empleado(
                codigo=codigo,
                nombre=nombre,
                apellido=apellido,
                ci=ci,
                cargo_id=int(cargo_id),
                salario_base=Decimal(salario_base),
                fecha_ingreso=fecha_ingreso,
                email=request.form.get('email'),
                telefono=request.form.get('telefono'),
                direccion=request.form.get('direccion'),
                nacionalidad=request.form.get('nacionalidad'),
                ips_numero=request.form.get('ips_numero'),
                sexo=request.form.get('sexo'),
                fecha_nacimiento=request.form.get('fecha_nacimiento') and datetime.strptime(request.form.get('fecha_nacimiento'), '%Y-%m-%d').date()
            )
            
            db.session.add(empleado)
            db.session.commit()
            
            registrar_operacion_crud(
                current_user, 'empleados', 'CREATE', 'empleados', 
                empleado.id, {'codigo': codigo, 'nombre': nombre, 'apellido': apellido}
            )
            
            flash(f'Empleado {empleado.nombre_completo} creado exitosamente', 'success')
            return redirect(url_for('rrhh.listar_empleados'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear empleado: {str(e)}', 'danger')
    
    return render_template('rrhh/crear_empleado.html', cargos=cargos)

@rrhh_bp.route('/empleados/<int:empleado_id>/editar', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def editar_empleado(empleado_id):
    """Editar empleado"""
    empleado = Empleado.query.get_or_404(empleado_id)
    cargos = Cargo.query.all()
    
    if request.method == 'POST':
        try:
            empleado.nombre = request.form.get('nombre')
            empleado.apellido = request.form.get('apellido')
            empleado.cargo_id = int(request.form.get('cargo_id'))
            empleado.salario_base = Decimal(request.form.get('salario_base'))
            empleado.email = request.form.get('email')
            empleado.telefono = request.form.get('telefono')
            empleado.direccion = request.form.get('direccion')
            empleado.nacionalidad = request.form.get('nacionalidad')
            empleado.ips_numero = request.form.get('ips_numero')
            empleado.motivo_retiro = request.form.get('motivo_retiro')
            empleado.sexo = request.form.get('sexo')
            empleado.estado = EstadoEmpleadoEnum[request.form.get('estado')]
            
            fecha_nacimiento_str = request.form.get('fecha_nacimiento')
            if fecha_nacimiento_str:
                empleado.fecha_nacimiento = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
            
            db.session.commit()
            
            registrar_operacion_crud(
                current_user, 'empleados', 'UPDATE', 'empleados', 
                empleado_id, {'cambios': 'Datos del empleado actualizados'}
            )
            
            flash(f'Empleado {empleado.nombre_completo} actualizado exitosamente', 'success')
            return redirect(url_for('rrhh.listar_empleados'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar empleado: {str(e)}', 'danger')
    
    return render_template('rrhh/editar_empleado.html', empleado=empleado, cargos=cargos)


@rrhh_bp.route('/api/empleados/<int:empleado_id>/general', methods=['GET'], endpoint='api_empleado_general_v2')
@login_required
def api_empleado_general_v2(empleado_id):
    """API usada por el perfil del empleado: KPIs y resumen de hoy, incluyendo Ingresos Extras."""
    empleado = Empleado.query.get_or_404(empleado_id)

    # Asistencia mes (conteo de días presentes en el mes actual)
    hoy = date.today()
    asistencia_mes = db.session.query(func.count(Asistencia.id)).filter(
        Asistencia.empleado_id == empleado.id,
        func.extract('month', Asistencia.fecha) == hoy.month,
        func.extract('year', Asistencia.fecha) == hoy.year,
        Asistencia.presente == True
    ).scalar() or 0

    # Vacaciones pendientes: sumar días pendientes para el año actual (más representativo)
    vacaciones_pendientes = 0
    if 'Vacacion' in globals():
        try:
            vacaciones_pendientes = db.session.query(func.coalesce(func.sum(Vacacion.dias_pendientes), 0)).filter(
                Vacacion.empleado_id == empleado.id,
                Vacacion.año == hoy.year
            ).scalar() or 0
            vacaciones_pendientes = int(vacaciones_pendientes)
        except Exception:
            # fallback al conteo de solicitudes pendientes si hay algún problema
            vacaciones_pendientes = Vacacion.query.filter_by(empleado_id=empleado.id, estado=EstadoVacacionEnum.PENDIENTE).count()

    # Sanciones activas
    sanciones_activas = Sancion.query.filter_by(empleado_id=empleado.id).count()

    # Permisos usados (conteo de permisos en el mes)
    permisos_usados = Permiso.query.filter(
        Permiso.empleado_id == empleado.id,
        func.extract('month', Permiso.fecha_creacion) == hoy.month,
        func.extract('year', Permiso.fecha_creacion) == hoy.year
    ).count()

    # Resumen hoy
    resumen_hoy = _resumir_dia_asistencias(empleado.id, hoy)

    # Ingresos extras resumen
    ingresos_pending_count = IngresoExtra.query.filter_by(empleado_id=empleado.id, estado='PENDIENTE').count()
    ingresos_pending_total = db.session.query(func.coalesce(func.sum(IngresoExtra.monto), 0)).filter(
        IngresoExtra.empleado_id == empleado.id,
        IngresoExtra.estado == 'PENDIENTE'
    ).scalar() or 0

    ultimo_ingreso = IngresoExtra.query.filter_by(empleado_id=empleado.id).order_by(IngresoExtra.id.desc()).first()
    ultimo = None
    if ultimo_ingreso:
        ultimo = {
            'id': ultimo_ingreso.id,
            'tipo': getattr(ultimo_ingreso, 'tipo', None),
            'monto': float(ultimo_ingreso.monto) if getattr(ultimo_ingreso, 'monto', None) is not None else None,
            'mes': getattr(ultimo_ingreso, 'mes', None),
            'año': getattr(ultimo_ingreso, 'año', None),
            'estado': getattr(ultimo_ingreso, 'estado', None),
            'justificativo': getattr(ultimo_ingreso, 'justificativo_archivo', None)
        }

    # calcular antiguedad en años (texto)
    antig = '-'
    if empleado.fecha_ingreso:
        años = (date.today() - empleado.fecha_ingreso).days // 365
        antig = f"{años} años"

    data = {
        'asistencia_mes': int(asistencia_mes),
        'vacaciones_pendientes': int(vacaciones_pendientes) if vacaciones_pendientes is not None else 0,
        'sanciones_activas': int(sanciones_activas),
        'permisos_usados': int(permisos_usados),
        'salario_base': float(empleado.salario_base) if empleado.salario_base is not None else 0,
        'fecha_ingreso': empleado.fecha_ingreso.strftime('%Y-%m-%d') if empleado.fecha_ingreso else None,
        'antiguedad': antig,
        'estado': 'Activo' if getattr(empleado.estado, 'name', '').upper() == 'ACTIVO' else (empleado.estado.value if hasattr(empleado.estado, 'value') else str(empleado.estado)),
        'email': empleado.email,
        'telefono': empleado.telefono,
        'resumen_hoy': resumen_hoy,
        'ingresos_extras': {
            'pendientes_count': int(ingresos_pending_count),
            'pendientes_total': float(ingresos_pending_total),
            'ultimo': ultimo
        }
    }

    return jsonify(data)

@rrhh_bp.route('/empleados/<int:empleado_id>/eliminar', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def eliminar_empleado(empleado_id):
    """Eliminar empleado (solo RRHH)"""
    empleado = Empleado.query.get_or_404(empleado_id)
    
    try:
        codigo = empleado.codigo
        db.session.delete(empleado)
        db.session.commit()
        
        registrar_operacion_crud(
            current_user, 'empleados', 'DELETE', 'empleados', 
            empleado_id, {'codigo': codigo}
        )
        
        flash('Empleado eliminado exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar empleado: {str(e)}', 'danger')
    
    return redirect(url_for('rrhh.listar_empleados'))


# ===================== PLANILLAS MTESS =====================
@rrhh_bp.route('/planillas/mtess', methods=['GET'])
@login_required
def planillas_mtess():
    hoy = date.today()
    return render_template('planillas/mtess.html', current_year=hoy.year)


@rrhh_bp.route('/planillas/mtess/download', methods=['GET'])
@login_required
def planillas_mtess_download():
    tipo = request.args.get('tipo') or 'personal'
    anio = int(request.args.get('anio') or date.today().year)
    mes = request.args.get('mes')
    mes = int(mes) if mes else None

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    detalle_bitacora = {'tipo': tipo, 'anio': anio, 'mes': mes}

    if tipo == 'personal':
        ws = wb.create_sheet('Personal')
        headers = ['Nombre', 'Apellido', 'Cédula', 'Fecha Ingreso', 'Fecha Salida', 'Cargo', 'Sexo', 'Edad', 'Nacionalidad', 'Motivo Salida']
        ws.append(headers)

        # empleados que tuvieron relación durante el año
        fin_anio = date(anio, 12, 31)
        empleados = Empleado.query.filter(Empleado.fecha_ingreso <= fin_anio).all()
        for e in empleados:
            fecha_salida = e.fecha_retiro.strftime('%d/%m/%Y') if getattr(e, 'fecha_retiro', None) else ''
            fecha_ing = e.fecha_ingreso.strftime('%d/%m/%Y') if e.fecha_ingreso else ''
            edad = ''
            if getattr(e, 'fecha_nacimiento', None):
                try:
                    edad = str((date(anio,12,31) - e.fecha_nacimiento).days // 365)
                except Exception:
                    edad = ''
            nacionalidad = getattr(e, 'nacionalidad', '') if hasattr(e, 'nacionalidad') else ''
            motivo = ''
            sexo = (getattr(e, 'sexo') or '')
            ws.append([e.nombre, e.apellido, getattr(e, 'ci', ''), fecha_ing, fecha_salida, getattr(e.cargo, 'nombre', ''), sexo, edad, nacionalidad, motivo])

    elif tipo == 'sueldos':
        ws = wb.create_sheet('Sueldos')
        headers = ['Cédula', 'Nombre', 'Período', 'Salario Base', 'Ingresos Extras', 'Descuentos', 'Aporte IPS', 'Salario Neto', 'Días Trabajados']
        ws.append(headers)

        # filtrar liquidaciones por año/mes
        if mes:
            periodo = f"{anio}-{mes:02d}"
            liquidaciones = Liquidacion.query.filter(Liquidacion.periodo == periodo).all()
        else:
            liquidaciones = Liquidacion.query.filter(Liquidacion.periodo.like(f"{anio}-%")).all()

        for l in liquidaciones:
            emp = l.empleado
            ws.append([getattr(emp, 'ci', ''), emp.nombre_completo, l.periodo, float(l.salario_base or 0), float(l.ingresos_extras or 0), float(l.descuentos or 0), float(l.aporte_ips or 0), float(l.salario_neto or 0), int(l.dias_trabajados or 0)])

    else:  # resumen
        ws = wb.create_sheet('Resumen')
        headers = ['Métrica', 'Valor']
        ws.append(headers)

        fin_anio = date(anio,12,31)
        empleados = Empleado.query.filter(Empleado.fecha_ingreso <= fin_anio).all()
        total = len(empleados)
        hombres = sum(1 for e in empleados if (getattr(e, 'sexo') or '').upper() == 'M')
        mujeres = sum(1 for e in empleados if (getattr(e, 'sexo') or '').upper() == 'F')
        menores = 0
        for e in empleados:
            if getattr(e, 'fecha_nacimiento', None):
                edad = (date(anio,12,31) - e.fecha_nacimiento).days // 365
                if edad < 18:
                    menores += 1

        # por cargos (categorías)
        cargos = db.session.query(Cargo.nombre, func.count(Empleado.id)).join(Empleado).group_by(Cargo.nombre).all()
        extranjeros = 0

        ws.append(['Total trabajadores', total])
        ws.append(['Hombres', hombres])
        ws.append(['Mujeres', mujeres])
        ws.append(['Menores de edad', menores])
        for c_name, cnt in cargos:
            ws.append([f'Cargo: {c_name}', cnt])
        ws.append(['Empleados extranjeros (no disponible)', extranjeros])

    # Generar archivo en memoria
    out = IOBytes()
    wb.save(out)
    out.seek(0)

    filename = f"MTESS_{tipo}_{anio}{('_%02d' % mes) if mes else ''}.xlsx"

    # Registrar en bitácora
    try:
        registrar_bitacora(current_user, 'planillas', 'DOWNLOAD', 'planillas_mtess', None, json.dumps(detalle_bitacora))
    except Exception as e:
        print('Error bitacora planillas:', e)

    return send_file(out, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ===================== PLANILLAS IPS/REI =====================
@rrhh_bp.route('/planillas/ips-rei', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def planillas_ips_rei():
    """Vista para generar planilla IPS/REI (usa la empresa única del sistema)"""
    hoy = date.today()
    empresa = Empresa.query.first()
    return render_template('planillas/ips_rei.html', current_year=hoy.year, empresa=empresa)


@rrhh_bp.route('/planillas/ips-rei/download', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def planillas_ips_rei_download():
    """Descargar planilla IPS/REI en formato Excel exacto"""
    from ..ips_utils import generar_fila_planilla_ips

    mes = request.args.get('mes', type=int)
    anio = request.args.get('anio', type=int)

    if not all([mes, anio]):
        flash('Mes y año son requeridos', 'danger')
        return redirect(url_for('rrhh.planillas_ips_rei'))

    # Obtener la empresa configurada
    empresa = Empresa.query.first()
    if not empresa:
        flash('No hay empresa configurada en el sistema', 'danger')
        return redirect(url_for('rrhh.planillas_ips_rei'))

    # Validar numero patronal
    if not empresa.numero_patronal:
        flash(f'⚠️ La empresa "{empresa.nombre}" no tiene número patronal configurado. Por favor, agregalo antes de generar la planilla.', 'danger')
        return redirect(url_for('rrhh.planillas_ips_rei'))

    # Periodo
    periodo = f"{anio}-{mes:02d}"

    # Obtener liquidaciones del mes SOLO para empleados ACTIVOS
    liquidaciones = Liquidacion.query.filter(
        Liquidacion.periodo == periodo,
        Liquidacion.empleado.has(Empleado.estado == EstadoEmpleadoEnum.ACTIVO)
    ).all()

    if not liquidaciones:
        flash(f'No hay liquidaciones de empleados ACTIVOS para {periodo}', 'warning')
        return redirect(url_for('rrhh.planillas_ips_rei'))

    # Validar que todos los empleados tengan ips_numero
    empleados_sin_ips = [liq for liq in liquidaciones if not liq.empleado.ips_numero]
    if empleados_sin_ips:
        nombres_sin_ips = ', '.join([e.empleado.nombre_completo for e in empleados_sin_ips])
        flash(f'⚠️ Los siguientes empleados no tienen número IPS asignado: {nombres_sin_ips}', 'warning')
        # Continuamos de todas formas (pero mostramos la advertencia)

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'REI'

    # Headers exactos según formato IPS/REI
    headers = [
        'Numero Patronal',
        'RUC Empresa',
        'Razon Social',
        'Numero Hoja',
        'Cedula',
        'Numero Asegurado',
        'Apellidos',
        'Nombres',
        'Dias Trabajados',
        'Salario Imponible',
        'Categoria',
        'Codigo Situacion',
        'Total Trabajadores (Hoja)',
        'Total Salario Imponible (Hoja)',
        'Aporte Empleado',
        'Aporte Empleador',
        'Total Aporte'
    ]
    ws.append(headers)

    # Agregar filas de datos
    total_trabajadores = len(liquidaciones)
    total_salario_imponible = sum(float(l.salario_base or 0) for l in liquidaciones)

    numero_hoja = 1
    filas_por_hoja = 50
    fila_actual = 0

    for liq in liquidaciones:
        # Si superamos filas_por_hoja, crear nueva hoja
        if fila_actual >= filas_por_hoja:
            numero_hoja += 1
            ws = wb.create_sheet(f'REI_{numero_hoja}')
            ws.append(headers)
            fila_actual = 0

        fila = generar_fila_planilla_ips(liq.empleado, liq, empresa)
        fila['numero_hoja'] = numero_hoja

        ws.append([
            fila['numero_patronal'],
            fila['ruc_empresa'],
            fila['razon_social'],
            fila['numero_hoja'],
            fila['cedula'],
            fila['numero_asegurado'],
            fila['apellidos'],
            fila['nombres'],
            fila['dias_trabajados'],
            fila['salario_imponible'],
            fila['categoria'],
            fila['codigo_situacion'],
            total_trabajadores,
            total_salario_imponible,
            fila['aporte_empleado'],
            fila['aporte_empleador'],
            fila['total_aporte'],
        ])
        fila_actual += 1

    # Generar archivo en memoria
    out = IOBytes()
    wb.save(out)
    out.seek(0)


    filename = f"REI_{empresa.nombre.replace(' ', '_')}_{periodo}.xlsx"

    # Registrar en bitácora
    try:
        registrar_bitacora(
            current_user, 'planillas', 'DOWNLOAD', 'planillas_ips_rei',
            None, json.dumps({'periodo': periodo})
        )
    except Exception as e:
        print('Error bitacora planillas IPS/REI:', e)

    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ===================== EMPRESAS =====================
@rrhh_bp.route('/empresas', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def listar_empresas():
    """Lista todas las empresas"""
    registrar_bitacora(current_user, 'empresas', 'VIEW', 'empresas')
    empresas = Empresa.query.all()
    return render_template('rrhh/empresas.html', empresas=empresas)


@rrhh_bp.route('/empresas/<int:empresa_id>/editar', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def editar_empresa(empresa_id):
    """Editar empresa (número patronal, etc.)"""
    empresa = Empresa.query.get_or_404(empresa_id)
    
    if request.method == 'POST':
        try:
            empresa.nombre = request.form.get('nombre') or empresa.nombre
            empresa.ruc = request.form.get('ruc') or empresa.ruc
            empresa.razon_social = request.form.get('razon_social') or empresa.razon_social
            empresa.numero_patronal = request.form.get('numero_patronal')
            empresa.direccion = request.form.get('direccion')
            empresa.telefono = request.form.get('telefono')
            empresa.email = request.form.get('email')
            empresa.ciudad = request.form.get('ciudad')
            empresa.representante_legal = request.form.get('representante_legal')
            empresa.ci_representante = request.form.get('ci_representante')
            
            # Porcentajes IPS
            ips_emp_str = request.form.get('porcentaje_ips_empleado')
            ips_empr_str = request.form.get('porcentaje_ips_empleador')
            
            if ips_emp_str:
                try:
                    empresa.porcentaje_ips_empleado = Decimal(ips_emp_str)
                except:
                    pass
            
            if ips_empr_str:
                try:
                    empresa.porcentaje_ips_empleador = Decimal(ips_empr_str)
                except:
                    pass
            
            db.session.commit()
            
            registrar_operacion_crud(
                current_user, 'empresas', 'UPDATE', 'empresas',
                empresa_id, {
                    'nombre': empresa.nombre,
                    'numero_patronal': empresa.numero_patronal,
                    'cambios': 'Datos de empresa actualizados'
                }
            )
            
            flash(f'Empresa "{empresa.nombre}" actualizada exitosamente', 'success')
            return redirect(url_for('rrhh.listar_empresas'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar empresa: {str(e)}', 'danger')
    
    return render_template('rrhh/editar_empresa.html', empresa=empresa)


# ==================== CARGOS ====================
@rrhh_bp.route('/cargos', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def listar_cargos():
    """Lista todos los cargos"""
    registrar_bitacora(current_user, 'cargos', 'VIEW', 'cargos')
    cargos = Cargo.query.all()
    return render_template('rrhh/cargos.html', cargos=cargos)

@rrhh_bp.route('/cargos/crear', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def crear_cargo():
    """Crear nuevo cargo"""
    if request.method == 'POST':
        try:
            nombre = request.form.get('nombre')
            
            if Cargo.query.filter_by(nombre=nombre).first():
                flash('El cargo ya existe', 'danger')
                return redirect(url_for('rrhh.crear_cargo'))
            
            cargo = Cargo(
                nombre=nombre,
                descripcion=request.form.get('descripcion'),
                salario_base=Decimal(request.form.get('salario_base'))
            )
            
            db.session.add(cargo)
            db.session.commit()
            
            registrar_operacion_crud(
                current_user, 'cargos', 'CREATE', 'cargos', 
                cargo.id, {'nombre': nombre}
            )
            
            flash(f'Cargo {cargo.nombre} creado exitosamente', 'success')
            return redirect(url_for('rrhh.listar_cargos'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear cargo: {str(e)}', 'danger')
    
    return render_template('rrhh/crear_cargo.html')

@rrhh_bp.route('/cargos/<int:cargo_id>/editar', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def editar_cargo(cargo_id):
    """Editar cargo"""
    cargo = Cargo.query.get_or_404(cargo_id)
    
    if request.method == 'POST':
        try:
            cargo.nombre = request.form.get('nombre')
            cargo.descripcion = request.form.get('descripcion')
            cargo.salario_base = Decimal(request.form.get('salario_base'))
            
            db.session.commit()
            
            registrar_operacion_crud(
                current_user, 'cargos', 'UPDATE', 'cargos', 
                cargo_id, {'cambios': 'Datos del cargo actualizados'}
            )
            
            flash(f'Cargo {cargo.nombre} actualizado exitosamente', 'success')
            return redirect(url_for('rrhh.listar_cargos'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar cargo: {str(e)}', 'danger')
    
    return render_template('rrhh/editar_cargo.html', cargo=cargo)

@rrhh_bp.route('/cargos/<int:cargo_id>/eliminar', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def eliminar_cargo(cargo_id):
    """Eliminar cargo"""
    cargo = Cargo.query.get_or_404(cargo_id)
    
    # Verificar si hay empleados con este cargo
    if cargo.empleados:
        flash('No se puede eliminar un cargo que tiene empleados asignados', 'danger')
        return redirect(url_for('rrhh.listar_cargos'))
    
    try:
        nombre = cargo.nombre
        db.session.delete(cargo)
        db.session.commit()
        
        registrar_operacion_crud(
            current_user, 'cargos', 'DELETE', 'cargos', 
            cargo_id, {'nombre': nombre}
        )
        
        flash('Cargo eliminado exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar cargo: {str(e)}', 'danger')
    
    return redirect(url_for('rrhh.listar_cargos'))

# ==================== EMPRESA ====================
@rrhh_bp.route('/empresa', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def ver_empresa():
    """Ver y editar datos de la empresa (configuración única)"""
    empresa = Empresa.query.first()
    if not empresa:
        empresa = Empresa(nombre='Mi Empresa')
        db.session.add(empresa)
        db.session.commit()

    if request.method == 'POST':
        try:
            empresa.nombre = request.form.get('nombre') or empresa.nombre
            empresa.ruc = request.form.get('ruc') or empresa.ruc
            empresa.razon_social = request.form.get('razon_social') or empresa.razon_social
            # Número patronal editable desde la UI
            empresa.numero_patronal = request.form.get('numero_patronal') or empresa.numero_patronal
            empresa.direccion = request.form.get('direccion') or empresa.direccion
            empresa.telefono = request.form.get('telefono') or empresa.telefono
            empresa.email = request.form.get('email') or empresa.email
            empresa.ciudad = request.form.get('ciudad') or empresa.ciudad
            empresa.representante_legal = request.form.get('representante_legal') or empresa.representante_legal
            empresa.ci_representante = request.form.get('ci_representante') or empresa.ci_representante

            # Porcentajes IPS (opcional)
            ips_emp_str = request.form.get('porcentaje_ips_empleado')
            ips_empr_str = request.form.get('porcentaje_ips_empleador')
            if ips_emp_str:
                try:
                    empresa.porcentaje_ips_empleado = Decimal(ips_emp_str)
                except:
                    pass
            if ips_empr_str:
                try:
                    empresa.porcentaje_ips_empleador = Decimal(ips_empr_str)
                except:
                    pass

            # dias_habiles_mes
            if request.form.get('dias_habiles_mes'):
                try:
                    empresa.dias_habiles_mes = int(request.form.get('dias_habiles_mes'))
                except:
                    pass

            # Logo upload (optional)
            if 'logo' in request.files:
                file = request.files['logo']
                if file and allowed_file(file.filename):
                    filename = secure_filename(f"logo_{int(datetime.utcnow().timestamp())}.{file.filename.rsplit('.', 1)[1].lower()}")
                    logo_path = os.path.join(UPLOADS_ROOT, 'empresa')
                    os.makedirs(logo_path, exist_ok=True)
                    file.save(os.path.join(logo_path, filename))
                    empresa.logo_path = f"empresa/{filename}"

            db.session.commit()

            registrar_operacion_crud(
                current_user, 'empresa', 'UPDATE', 'empresas', 
                empresa.id, {'nombre': empresa.nombre, 'numero_patronal': empresa.numero_patronal}
            )

            flash('Datos de la empresa actualizados exitosamente', 'success')
            return redirect(url_for('rrhh.ver_empresa'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar empresa: {str(e)}', 'danger')

    registrar_bitacora(current_user, 'empresa', 'VIEW', 'empresas')
    return render_template('rrhh/empresa.html', empresa=empresa)

# ==================== ASISTENCIA ====================
@rrhh_bp.route('/asistencia', methods=['GET'])
@login_required
def ver_asistencia():
    """Vista de asistencia para registrar entrada/salida"""
    registrar_bitacora(current_user, 'asistencia', 'VIEW', 'asistencias')
    
    pagina = request.args.get('page', 1, type=int)
    fecha_filtro = request.args.get('fecha', str(date.today()))
    
    try:
        fecha_date = datetime.strptime(fecha_filtro, '%Y-%m-%d').date()
    except:
        fecha_date = date.today()
    
    asistencias = Asistencia.query.filter_by(fecha=fecha_date).paginate(page=pagina, per_page=20)
    
    return render_template('rrhh/asistencia.html', asistencias=asistencias, fecha=fecha_filtro)


def _parse_time(cfg_key, default='08:30'):
    s = current_app.config.get(cfg_key, default)
    try:
        return datetime.strptime(s, '%H:%M').time()
    except Exception:
        return datetime.strptime(default, '%H:%M').time()


def _resumir_dia_asistencias(empleado_id, dia_date):
    """Resumir los eventos del día y devolver dict con hora_entrada, hora_salida, presente, observaciones"""
    eventos = AsistenciaEvento.query.filter(
        AsistenciaEvento.empleado_id == empleado_id,
        func.date(AsistenciaEvento.ts) == dia_date
    ).order_by(AsistenciaEvento.ts).all()

    on_time = _parse_time('ASISTENCIA_ON_TIME', '08:30')
    tolerancia = _parse_time('ASISTENCIA_TOLERANCE', '08:50')
    corte_manana = _parse_time('ASISTENCIA_CORTE_MANANA', '12:00')
    inicio_tarde = _parse_time('ASISTENCIA_INICIO_TARDE', '13:30')
    salida_final = _parse_time('ASISTENCIA_SALIDA_FINAL', '16:00')

    primera_entrada = None
    ultima_salida = None
    observaciones = []

    # Encontrar primera entrada y última salida
    for ev in eventos:
        if ev.tipo == 'in' and primera_entrada is None:
            primera_entrada = ev.ts
        if ev.tipo == 'out':
            ultima_salida = ev.ts

    # Clasificar llegada
    if primera_entrada:
        t = primera_entrada.time()
        if t <= on_time:
            observaciones.append('A tiempo')
        elif t <= tolerancia:
            observaciones.append('Tolerancia')
        elif t < corte_manana:
            observaciones.append('Llegada tardía')
        else:
            # Entrada después del corte de mañana -> considerarla llegada tarde o sólo tarde
            if t >= inicio_tarde:
                observaciones.append('Solo tarde')
            else:
                observaciones.append('Llegada tardía')
    else:
        # No hay entrada por la mañana
        # Si existe entrada por la tarde, marcar solo tarde
        entrada_tarde = next((ev for ev in eventos if ev.tipo == 'in' and ev.ts.time() >= inicio_tarde), None)
        if entrada_tarde:
            observaciones.append('Ausente mañana; Solo tarde')
            primera_entrada = entrada_tarde.ts if hasattr(entrada_tarde, 'ts') else entrada_tarde.ts
        else:
            observaciones.append('Ausencia')

    # Detectar salida al almuerzo (primer out cercano a medio día)
    lunch_out = next((ev for ev in eventos if ev.tipo == 'out' and ev.ts.time() >= (datetime.strptime('11:30','%H:%M').time()) and ev.ts.time() <= (datetime.strptime('13:30','%H:%M').time())), None)
    if lunch_out:
        observaciones.append('Salida al almuerzo')
        # buscar entrada después del almuerzo
        lunch_in = next((ev for ev in eventos if ev.tipo == 'in' and ev.ts > lunch_out.ts), None)
        if lunch_in:
            observaciones.append('Entrada del almuerzo')

    # Clasificar salida final
    if ultima_salida:
        if ultima_salida.time() >= salida_final:
            observaciones.append('Salida final')
        else:
            observaciones.append('Salida temprana')

    resumen = {
        'hora_entrada': primera_entrada.time().strftime('%H:%M:%S') if primera_entrada else None,
        'hora_salida': ultima_salida.time().strftime('%H:%M:%S') if ultima_salida else None,
        'presente': True if eventos else False,
        'observaciones': '; '.join(observaciones) if observaciones else None
    }
    return resumen


def cerrar_asistencias_automatico(fecha_cierre=None):
    """
    Cierra automáticamente las asistencias del día especificado.
    Genera registros para empleados sin marcación, distinguiendo:
    - Vacaciones aprobadas
    - Permisos aprobados
    - Ausencias injustificadas
    
    Args:
        fecha_cierre: Fecha a cerrar (date). Si None, usa fecha actual.
    
    Returns:
        dict con estadísticas de procesamiento
    """
    from app.models import Empleado, EstadoEmpleadoEnum, Vacacion, EstadoVacacionEnum, Permiso, EstadoPermisoEnum
    
    if fecha_cierre is None:
        fecha_cierre = date.today()
    
    # Solo procesar días laborales (lunes a sábado)
    if fecha_cierre.weekday() == 6:  # Domingo = 6
        return {
            'procesados': 0,
            'mensaje': f'Domingo {fecha_cierre} no es día laboral (omitido)',
            'vacaciones': 0,
            'permisos': 0,
            'ausencias': 0
        }
    
    # Obtener empleados activos
    empleados_activos = Empleado.query.filter_by(
        estado=EstadoEmpleadoEnum.ACTIVO
    ).all()
    
    stats = {
        'procesados': 0,
        'vacaciones': 0,
        'permisos': 0,
        'ausencias': 0,
        'ya_registrados': 0
    }
    
    for empleado in empleados_activos:
        # Saltar si ingresó después de la fecha de cierre
        if empleado.fecha_ingreso and empleado.fecha_ingreso > fecha_cierre:
            continue
        
        # Saltar si salió antes de la fecha de cierre
        if empleado.fecha_salida and empleado.fecha_salida < fecha_cierre:
            continue
        
        # Verificar si ya tiene asistencia registrada
        asistencia_existe = Asistencia.query.filter_by(
            empleado_id=empleado.id,
            fecha=fecha_cierre
        ).first()
        
        if asistencia_existe:
            stats['ya_registrados'] += 1
            continue
        
        # 1. Verificar vacaciones aprobadas
        vacacion = Vacacion.query.filter(
            Vacacion.empleado_id == empleado.id,
            Vacacion.estado == EstadoVacacionEnum.APROBADA,
            Vacacion.fecha_inicio_solicitud <= fecha_cierre,
            Vacacion.fecha_fin_solicitud >= fecha_cierre
        ).first()
        
        if vacacion:
            asistencia = Asistencia(
                empleado_id=empleado.id,
                fecha=fecha_cierre,
                presente=True,  # Vacaciones cuentan como día trabajado
                observaciones='Vacaciones (auto-generado)',
                justificacion_estado=None  # No requiere justificación
            )
            db.session.add(asistencia)
            stats['vacaciones'] += 1
            stats['procesados'] += 1
            continue
        
        # 2. Verificar permisos aprobados
        permiso = Permiso.query.filter(
            Permiso.empleado_id == empleado.id,
            Permiso.estado.in_([EstadoPermisoEnum.APROBADO, EstadoPermisoEnum.COMPLETADO]),
            Permiso.fecha_inicio <= fecha_cierre,
            Permiso.fecha_fin >= fecha_cierre
        ).first()
        
        if permiso:
            asistencia = Asistencia(
                empleado_id=empleado.id,
                fecha=fecha_cierre,
                presente=True,  # Permiso cuenta como presente (si es remunerado)
                observaciones=f'Permiso: {permiso.motivo} (auto-generado)',
                justificacion_estado=None  # No requiere justificación
            )
            db.session.add(asistencia)
            stats['permisos'] += 1
            stats['procesados'] += 1
            continue
        
        # 3. Sin marcación ni justificación = Ausencia pendiente de justificar
        asistencia = Asistencia(
            empleado_id=empleado.id,
            fecha=fecha_cierre,
            presente=False,
            observaciones='Ausencia sin marcación (auto-generado)',
            justificacion_estado='PENDIENTE'  # Requiere decisión de RRHH
        )
        db.session.add(asistencia)
        stats['ausencias'] += 1
        stats['procesados'] += 1
    
    try:
        db.session.commit()
        stats['mensaje'] = f'Asistencias del {fecha_cierre} cerradas exitosamente'
    except Exception as e:
        db.session.rollback()
        stats['mensaje'] = f'Error al cerrar asistencias: {str(e)}'
        stats['error'] = True
    
    return stats


@rrhh_bp.route('/asistencia/registrar', methods=['POST'])
@login_required
def registrar_asistencia():
    """Registra un evento de asistencia (punch) y actualiza el resumen diario."""
    try:
        codigo = request.json.get('codigo', '').strip().upper()
        if not codigo:
            return jsonify({'success': False, 'message': 'Código de empleado requerido'}), 400

        empleado = Empleado.query.filter_by(codigo=codigo).first()
        if not empleado:
            return jsonify({'success': False, 'message': f'Empleado con código {codigo} no encontrado'}), 404

        if empleado.estado.name != 'ACTIVO':
            return jsonify({'success': False, 'message': f'El empleado está {empleado.estado.value}'}), 403

        ahora = datetime.now()
        hoy = date.today()
        hora_actual = ahora.time()
        
        # Hora límite de cierre (17:30)
        hora_cierre = datetime.strptime('17:30', '%H:%M').time()

        # Inferir tipo si no se pasa explícitamente
        tipo = request.json.get('tipo')
        eventos_count = AsistenciaEvento.query.filter(
            AsistenciaEvento.empleado_id == empleado.id,
            func.date(AsistenciaEvento.ts) == hoy
        ).count()
        if tipo not in ('in', 'out'):
            tipo = 'in' if eventos_count % 2 == 0 else 'out'
        
        # Validar: después de 17:30 solo se permite salida, NO entrada
        if hora_actual > hora_cierre and tipo == 'in':
            return jsonify({
                'success': False, 
                'message': f'No se permite marcar entrada después de las 17:30. Hora actual: {hora_actual.strftime("%H:%M")}'
            }), 403

        evento = AsistenciaEvento(
            empleado_id=empleado.id,
            ts=ahora,
            tipo=tipo,
            origen=request.json.get('origen', 'web'),
            detalles=json.dumps({'ip': request.remote_addr, 'user_agent': request.headers.get('User-Agent')})
        )
        db.session.add(evento)
        db.session.commit()

        # Recalcular y almacenar resumen diario
        resumen = _resumir_dia_asistencias(empleado.id, hoy)

        asistencia = Asistencia.query.filter_by(empleado_id=empleado.id, fecha=hoy).first()
        if not asistencia:
            asistencia = Asistencia(empleado_id=empleado.id, fecha=hoy)
            db.session.add(asistencia)

        asistencia.hora_entrada = resumen.get('hora_entrada')
        asistencia.hora_salida = resumen.get('hora_salida')
        asistencia.presente = resumen.get('presente', True)
        asistencia.observaciones = resumen.get('observaciones')
        db.session.commit()


        registrar_operacion_crud(current_user, 'asistencia', 'CREATE', 'asistencias', asistencia.id, {'empleado_codigo': codigo, 'evento_id': evento.id, 'tipo_evento': tipo})

        return jsonify({'success': True, 'message': f'Evento {tipo} registrado para {empleado.nombre_completo}', 'tipo': tipo, 'hora': ahora.strftime('%H:%M:%S'), 'resumen': resumen})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@rrhh_bp.route('/asistencia/<int:asistencia_id>/editar', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def editar_asistencia(asistencia_id):
    """Editar asistencia"""
    asistencia = Asistencia.query.get_or_404(asistencia_id)
    
    try:
        hora_entrada = request.form.get('hora_entrada')
        hora_salida = request.form.get('hora_salida')
        
        if hora_entrada:
            asistencia.hora_entrada = datetime.strptime(hora_entrada, '%H:%M').time()
        
        if hora_salida:
            asistencia.hora_salida = datetime.strptime(hora_salida, '%H:%M').time()
        
        asistencia.observaciones = request.form.get('observaciones')
        asistencia.presente = request.form.get('presente') == 'on'
        
        db.session.commit()
        
        registrar_operacion_crud(
            current_user, 'asistencia', 'UPDATE', 'asistencias',
            asistencia_id, {'cambios': 'Asistencia actualizada'}
        )
        
        flash('Asistencia actualizada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar asistencia: {str(e)}', 'danger')
    
    fecha = asistencia.fecha.strftime('%Y-%m-%d')
    return redirect(url_for('rrhh.ver_asistencia', fecha=fecha))


@rrhh_bp.route('/asistencia/cerrar_dia', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def cerrar_dia_asistencias():
    """Cierra manualmente las asistencias de un día específico"""
    try:
        fecha_str = request.form.get('fecha', str(date.today()))
        fecha_cierre = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        
        stats = cerrar_asistencias_automatico(fecha_cierre)
        
        if stats.get('error'):
            flash(stats['mensaje'], 'danger')
        else:
            flash(f"✅ {stats['mensaje']}: {stats['procesados']} registros generados "
                  f"({stats['vacaciones']} vacaciones, {stats['permisos']} permisos, "
                  f"{stats['ausencias']} ausencias pendientes)", 'success')
        
        registrar_bitacora(current_user, 'asistencia', 'CIERRE', 'asistencias', 
                          {'fecha': fecha_str, 'stats': stats})
        
    except Exception as e:
        flash(f'Error al cerrar asistencias: {str(e)}', 'danger')
    
    return redirect(url_for('rrhh.dashboard'))


@rrhh_bp.route('/asistencia/<int:asistencia_id>/justificar', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def justificar_ausencia(asistencia_id):
    """Marca una ausencia como justificada con nota de RRHH"""
    asistencia = Asistencia.query.get_or_404(asistencia_id)
    
    try:
        nota = request.form.get('justificacion_nota', '').strip()
        
        if not nota:
            return jsonify({'success': False, 'message': 'Debe proporcionar un motivo'}), 400
        
        asistencia.justificacion_estado = 'JUSTIFICADO'
        asistencia.justificacion_nota = nota
        asistencia.justificacion_fecha = datetime.now()
        asistencia.justificacion_por = current_user.id
        asistencia.observaciones = f'Ausencia justificada: {nota}'
        
        db.session.commit()
        
        registrar_operacion_crud(
            current_user, 'asistencia', 'UPDATE', 'asistencias',
            asistencia_id, {'accion': 'justificar', 'nota': nota}
        )
        
        return jsonify({
            'success': True, 
            'message': f'Ausencia de {asistencia.empleado.nombre_completo} justificada. Se descontará 1 día del salario.'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@rrhh_bp.route('/asistencia/<int:asistencia_id>/no_justificar', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def no_justificar_ausencia(asistencia_id):
    """Marca una ausencia como NO justificada (injustificada)"""
    asistencia = Asistencia.query.get_or_404(asistencia_id)
    
    try:
        nota = request.form.get('justificacion_nota', 'Sin motivo válido').strip()
        
        asistencia.justificacion_estado = 'INJUSTIFICADO'
        asistencia.justificacion_nota = nota
        asistencia.justificacion_fecha = datetime.now()
        asistencia.justificacion_por = current_user.id
        asistencia.observaciones = 'Ausencia injustificada'
        
        db.session.commit()
        
        registrar_operacion_crud(
            current_user, 'asistencia', 'UPDATE', 'asistencias',
            asistencia_id, {'accion': 'no_justificar', 'nota': nota}
        )
        
        return jsonify({
            'success': True, 
            'message': f'Ausencia de {asistencia.empleado.nombre_completo} marcada como injustificada. Se descontará 1 día del salario.'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@rrhh_bp.route('/permisos/empleado/<int:empleado_id>', methods=['GET'])
@login_required
def ver_historial_permisos(empleado_id):
    """Ver historial de permisos de un empleado, incluyendo observaciones"""
    registrar_bitacora(current_user, 'permisos', 'VIEW', 'permisos')

    empleado = Empleado.query.get_or_404(empleado_id)
    
    permisos = Permiso.query.filter_by(empleado_id=empleado_id).order_by(Permiso.fecha_creacion.desc()).all()
    pendientes = [p for p in permisos if p.estado.name == 'PENDIENTE']

    return render_template('rrhh/permiso_detalle.html', empleado=empleado, permisos=permisos, pendientes=pendientes)

@rrhh_bp.route('/permisos', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def listar_permisos():
    """Lista permisos"""
    registrar_bitacora(current_user, 'permisos', 'VIEW', 'permisos')
    
    pagina = request.args.get('page', 1, type=int)
    permisos = Permiso.query.paginate(page=pagina, per_page=10)
    
    return render_template('rrhh/permisos.html', permisos=permisos)

@rrhh_bp.route('/permisos/solicitar', methods=['GET', 'POST'])
@login_required
def solicitar_permiso():
    """Solicitar permiso"""
    empleados = Empleado.query.filter_by(estado=EstadoEmpleadoEnum.ACTIVO).all()
    
    if request.method == 'POST':
        try:
            fecha_inicio = datetime.strptime(request.form.get('fecha_inicio'), '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(request.form.get('fecha_fin'), '%Y-%m-%d').date()
            dias = (fecha_fin - fecha_inicio).days + 1

            # Validación mínima: exigir al menos 1 día
            if dias < 1:
                flash('El permiso debe incluir al menos un día', 'danger')
                return redirect(url_for('rrhh.solicitar_permiso'))
            # Determinar si el permiso es con goce salarial (checkbox en el form)
            con_goce_field = request.form.get('con_goce')
            con_goce = False
            if con_goce_field in ('on', 'true', '1', 'True'):
                con_goce = True

            permiso = Permiso(
                empleado_id=int(request.form.get('empleado_id')),
                tipo_permiso=request.form.get('tipo_permiso'),
                motivo=request.form.get('motivo'),
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                dias_solicitados=dias,
                observaciones=request.form.get('observaciones'),
                con_goce=con_goce
            )
            
            db.session.add(permiso)
            db.session.commit()
            
            registrar_operacion_crud(
                current_user, 'permisos', 'CREATE', 'permisos',
                permiso.id, {'empleado_id': permiso.empleado_id, 'tipo': permiso.tipo_permiso}
            )
            
            flash('Permiso solicitado exitosamente', 'success')
            return redirect(url_for('rrhh.listar_permisos'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al solicitar permiso: {str(e)}', 'danger')
    
    return render_template('rrhh/solicitar_permiso.html', empleados=empleados)

@rrhh_bp.route('/permisos/<int:permiso_id>/aprobar', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def aprobar_permiso(permiso_id):
    """Aprobar permiso"""
    permiso = Permiso.query.get_or_404(permiso_id)
    
    try:
        from ..models import EstadoPermisoEnum, Descuento
        from decimal import Decimal

        permiso.estado = EstadoPermisoEnum.APROBADO

        # Si el permiso es SIN GOCE (con_goce == False), crear Descuento(s) automáticos por mes
        # (dividir si cruza meses) solo si no hay descuentos activos ya creados para este permiso.
        if not permiso.con_goce:
            existing = Descuento.query.filter_by(origen_tipo='permiso', origen_id=permiso.id, activo=True).count()
            if existing:
                # ya existen descuentos activos para este permiso, no crear duplicados
                db.session.commit()
                registrar_operacion_crud(
                    current_user, 'permisos', 'UPDATE', 'permisos',
                    permiso_id, {'estado': 'APROBADO', 'info': 'descuentos_existentes'}
                )
                flash('Permiso aprobado (descuentos ya creados anteriormente)', 'success')
                return redirect(url_for('rrhh.listar_permisos'))
            empleado = permiso.empleado
            salario = Decimal(empleado.salario_base)

            fecha_start = permiso.fecha_inicio
            fecha_end = permiso.fecha_fin

            descuentos_creados = []

            # Iterar por segmentos por mes
            current = fecha_start
            while current <= fecha_end:
                last_day = calendar.monthrange(current.year, current.month)[1]
                end_of_month = date(current.year, current.month, last_day)
                segment_end = end_of_month if end_of_month <= fecha_end else fecha_end

                dias_segment = (segment_end - current).days + 1
                monto_segment = (salario / Decimal('30')) * Decimal(dias_segment)
                monto_segment = monto_segment.quantize(Decimal('0.01'))

                desc = Descuento(
                    empleado_id=empleado.id,
                    tipo='Permiso sin goce',
                    monto=monto_segment,
                    mes=current.month,
                    año=current.year,
                    descripcion=f'Descuento automático por permiso id {permiso.id} ({current.strftime("%Y-%m")})',
                    activo=True,
                    origen_tipo='permiso',
                    origen_id=permiso.id
                )
                db.session.add(desc)
                db.session.flush()
                descuentos_creados.append(desc.id)

                # mover al siguiente día después del segmento
                current = segment_end + timedelta(days=1)

            # Si se crearon descuentos, vincular el primero (para compatibilidad)
            if descuentos_creados:
                permiso.descuento_id = descuentos_creados[0]

        db.session.commit()

        registrar_operacion_crud(
            current_user, 'permisos', 'UPDATE', 'permisos',
            permiso_id, {'estado': 'APROBADO', 'descuento_id': permiso.descuento_id}
        )

        flash('Permiso aprobado exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al aprobar permiso: {str(e)}', 'danger')
    
    return redirect(url_for('rrhh.listar_permisos'))


@rrhh_bp.route('/ingresos-extras/<int:ingreso_id>/approve', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def aprobar_ingreso_extra(ingreso_id):
    ingreso = IngresoExtra.query.get_or_404(ingreso_id)
    try:
        ingreso.estado = 'APROBADO'
        ingreso.aprobado_por = current_user.id
        ingreso.fecha_aprobacion = datetime.utcnow()
        db.session.commit()
        registrar_operacion_crud(current_user, 'ingresos_extras', 'UPDATE', 'ingresos_extras', ingreso_id, {'estado': 'APROBADO'})
        return jsonify({'message': 'Ingreso extra aprobado'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@rrhh_bp.route('/ingresos-extras/<int:ingreso_id>/reject', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def rechazar_ingreso_extra(ingreso_id):
    ingreso = IngresoExtra.query.get_or_404(ingreso_id)
    try:
        ingreso.estado = 'RECHAZADO'
        ingreso.aprobado_por = current_user.id
        ingreso.fecha_aprobacion = datetime.utcnow()
        db.session.commit()
        registrar_operacion_crud(current_user, 'ingresos_extras', 'UPDATE', 'ingresos_extras', ingreso_id, {'estado': 'RECHAZADO'})
        return jsonify({'message': 'Ingreso extra rechazado'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@rrhh_bp.route('/permisos/<int:permiso_id>/anular', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def anular_permiso(permiso_id):
    """Anular un permiso (revertir descuento asociado si existe)."""
    permiso = Permiso.query.get_or_404(permiso_id)

    try:
        from ..models import EstadoPermisoEnum, Descuento

        permiso.estado = EstadoPermisoEnum.RECHAZADO

        # Marcar descuentos asociados inactivos (no borrarlos)
        descuentos = Descuento.query.filter_by(origen_tipo='permiso', origen_id=permiso.id, activo=True).all()
        for d in descuentos:
            d.activo = False

        if permiso.descuento_id:
            d = Descuento.query.get(permiso.descuento_id)
            if d:
                d.activo = False
            permiso.descuento_id = None

        db.session.commit()

        registrar_operacion_crud(
            current_user, 'permisos', 'UPDATE', 'permisos',
            permiso_id, {'accion': 'ANULAR', 'descuento_revertido': True}
        )

        flash('Permiso anulado y descuento revertido si existía', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al anular permiso: {str(e)}', 'danger')

    return redirect(url_for('rrhh.listar_permisos'))

@rrhh_bp.route('/permisos/<int:permiso_id>/rechazar', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def rechazar_permiso(permiso_id):
    """Rechazar permiso"""
    permiso = Permiso.query.get_or_404(permiso_id)
    
    try:
        from ..models import EstadoPermisoEnum, Descuento

        permiso.estado = EstadoPermisoEnum.RECHAZADO

        # Marcar como inactivos los descuentos asociados a este permiso (no borrarlos)
        descuentos = Descuento.query.filter_by(origen_tipo='permiso', origen_id=permiso.id, activo=True).all()
        for d in descuentos:
            d.activo = False

        # También si hay descuento_id directo, inactivarlo y limpiar el campo
        if permiso.descuento_id:
            d = Descuento.query.get(permiso.descuento_id)
            if d:
                d.activo = False
            permiso.descuento_id = None

        db.session.commit()

        registrar_operacion_crud(
            current_user, 'permisos', 'UPDATE', 'permisos',
            permiso_id, {'estado': 'RECHAZADO', 'descuento_revertido': True}
        )

        flash('Permiso rechazado', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al rechazar permiso: {str(e)}', 'danger')
    
    return redirect(url_for('rrhh.listar_permisos'))

# ==================== SANCIONES ====================
@rrhh_bp.route('/sanciones', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def listar_sanciones():
    """Lista sanciones"""
    registrar_bitacora(current_user, 'sanciones', 'VIEW', 'sanciones')
    
    pagina = request.args.get('page', 1, type=int)
    sanciones = Sancion.query.paginate(page=pagina, per_page=10)
    
    return render_template('rrhh/sanciones.html', sanciones=sanciones)

@rrhh_bp.route('/sanciones/crear', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def crear_sancion():
    """Crear sanción"""
    empleados = Empleado.query.filter_by(estado=EstadoEmpleadoEnum.ACTIVO).all()
    
    if request.method == 'POST':
        try:
            sancion = Sancion(
                empleado_id=int(request.form.get('empleado_id')),
                tipo_sancion=request.form.get('tipo_sancion'),
                motivo=request.form.get('motivo'),
                fecha=datetime.strptime(request.form.get('fecha'), '%Y-%m-%d').date(),
                monto=Decimal(request.form.get('monto', 0)),
                descripcion=request.form.get('descripcion')
            )
            
            db.session.add(sancion)
            db.session.commit()

            # Si la sanción es 'Suspensión' y se indicó días, crear descuentos automáticos
            try:
                dias_susp = int(request.form.get('dias_suspension', 0))
            except Exception:
                dias_susp = 0

            if sancion.tipo_sancion and 'suspension' in sancion.tipo_sancion.lower() and dias_susp > 0:
                from ..models import Descuento
                empleado = sancion.empleado
                salario = Decimal(empleado.salario_base)
                fecha_start = sancion.fecha
                fecha_end = fecha_start + timedelta(days=dias_susp-1)

                current = fecha_start
                while current <= fecha_end:
                    last_day = calendar.monthrange(current.year, current.month)[1]
                    end_of_month = date(current.year, current.month, last_day)
                    segment_end = end_of_month if end_of_month <= fecha_end else fecha_end

                    dias_segment = (segment_end - current).days + 1
                    monto_segment = (salario / Decimal('30')) * Decimal(dias_segment)
                    monto_segment = monto_segment.quantize(Decimal('0.01'))

                    desc = Descuento(
                        empleado_id=empleado.id,
                        tipo='Sancion - Suspensión',
                        monto=monto_segment,
                        mes=current.month,
                        año=current.year,
                        descripcion=f'Descuento por sanción id {sancion.id}',
                        activo=True,
                        origen_tipo='sancion',
                        origen_id=sancion.id
                    )
                    db.session.add(desc)
                    db.session.flush()

                    current = segment_end + timedelta(days=1)

                db.session.commit()
            
            registrar_operacion_crud(
                current_user, 'sanciones', 'CREATE', 'sanciones',
                sancion.id, {'empleado_id': sancion.empleado_id, 'tipo': sancion.tipo_sancion}
            )
            
            flash('Sanción registrada exitosamente', 'success')
            return redirect(url_for('rrhh.listar_sanciones'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar sanción: {str(e)}', 'danger')
    
    return render_template('rrhh/crear_sancion.html', empleados=empleados)

# ==================== LIQUIDACIONES ====================
@rrhh_bp.route('/liquidaciones', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def listar_liquidaciones():
    """Lista liquidaciones"""
    registrar_bitacora(current_user, 'liquidaciones', 'VIEW', 'liquidaciones')
    
    pagina = request.args.get('page', 1, type=int)
    periodo_filtro = request.args.get('periodo', f'{date.today().strftime("%Y-%m")}')
    
    liquidaciones = Liquidacion.query.filter_by(periodo=periodo_filtro).paginate(page=pagina, per_page=10)
    
    return render_template('rrhh/liquidaciones.html', liquidaciones=liquidaciones, periodo=periodo_filtro)


# ==================== INGRESOS EXTRAS / HORAS ====================
@rrhh_bp.route('/ingresos-extras', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def ver_ingresos_extras():
    """Página principal para administrar ingresos extras y horas extra"""
    registrar_bitacora(current_user, 'ingresos_extras', 'VIEW', 'ingresos_extras')
    # Pasar el módulo `date` para permitir usar `date.today()` en la plantilla
    # Usar plantilla limpia para evitar problemas si la original está corrupta
    return render_template('rrhh/ingresos_extras_clean.html', date=date)


@rrhh_bp.route('/api/empleados/<int:empleado_id>/ingresos-extras', methods=['GET'])
@login_required
def api_empleado_ingresos_extras(empleado_id):
    """Devuelve Ingresos Extra de un empleado (paginado)."""
    pagina = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)

    query = IngresoExtra.query.filter_by(empleado_id=empleado_id).order_by(IngresoExtra.id.desc())
    pag = query.paginate(page=pagina, per_page=per_page)

    items = []
    for ie in pag.items:
        items.append({
            'id': ie.id,
            'tipo': getattr(ie, 'tipo', None),
            'monto': float(ie.monto) if getattr(ie, 'monto', None) is not None else None,
            'mes': getattr(ie, 'mes', None),
            'año': getattr(ie, 'año', None),
            'estado': getattr(ie, 'estado', None),
            'justificativo': getattr(ie, 'justificativo_archivo', None),
            'fecha_creacion': ie.fecha_creacion.strftime('%Y-%m-%d') if getattr(ie, 'fecha_creacion', None) else None
        })

    return jsonify({
        'items': items,
        'page': pag.page,
        'pages': pag.pages,
        'per_page': pag.per_page,
        'total': pag.total
    })


@rrhh_bp.route('/ingresos-extras/create', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def crear_ingreso_extra():
    """Formulario para crear un Ingreso Extra (bonificación/comisión) manualmente."""
    # Asegurar carpeta de uploads
    os.makedirs(INGRESOS_UPLOAD_FOLDER, exist_ok=True)

    if request.method == 'POST':
        try:
            empleado_id = request.form.get('empleado_id')
            tipo = request.form.get('tipo')
            monto_raw = request.form.get('monto')
            periodo = request.form.get('periodo')  # YYYY-MM
            descripcion = request.form.get('descripcion')

            if not empleado_id or not tipo or not monto_raw or not periodo:
                flash('Faltan campos obligatorios', 'danger')
                return redirect(request.url)

            año, mes = map(int, periodo.split('-'))
            monto = Decimal(monto_raw.replace(',', ''))

            justificativo_filename = None
            if 'justificativo' in request.files:
                file = request.files['justificativo']
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    timestamp = int(datetime.utcnow().timestamp())
                    filename = f"ingreso_{empleado_id}_{timestamp}_{filename}"
                    filepath = os.path.join(INGRESOS_UPLOAD_FOLDER, filename)
                    file.save(filepath)
                    justificativo_filename = os.path.join('ingresos_extras', filename)

            ingreso = IngresoExtra(
                empleado_id=int(empleado_id),
                tipo=tipo,
                monto=monto,
                mes=mes,
                año=año,
                descripcion=descripcion,
                estado='PENDIENTE',
                creado_por=current_user.id,
                aplicado=False,
                justificativo_archivo=justificativo_filename
            )
            db.session.add(ingreso)
            db.session.commit()

            registrar_bitacora(current_user, 'ingresos_extras', 'CREATE', 'ingresos_extras', detalle=f'Ingreso extra creado para empleado {empleado_id} monto {monto}')
            flash('Ingreso extra creado correctamente (pendiente de aprobación)', 'success')
            return redirect(url_for('rrhh.ver_ingresos_extras'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error creando ingreso extra: {e}', 'danger')
            return redirect(request.url)

    # GET: mostrar formulario
    empleado = None
    empleado_id = request.args.get('empleado_id')
    if empleado_id:
        empleado = Empleado.query.get(empleado_id)

    periodo_default = date.today().strftime('%Y-%m')
    empleados = []
    if not empleado:
        empleados = Empleado.query.order_by(Empleado.apellido).all()
    return render_template('rrhh/ingreso_extra_form.html', empleado=empleado, periodo_default=periodo_default, empleados=empleados)


def _compute_horas_extra_for_month(empleado, year, month):
    """Detecta horas extra desde AsistenciaEvento y crea/actualiza registros en HorasExtra."""
    from calendar import monthrange
    empresa = Empresa.query.first()
    dias_habiles = empresa.dias_habiles_mes if empresa and empresa.dias_habiles_mes else 30
    horas_jornada = 8
    # fin jornada por defecto: 17:00
    fin_jornada_hour = 17

    primer_dia = date(year, month, 1)
    ultimo_dia = date(year, month, monthrange(year, month)[1])

    # obtener eventos out por día
    curr = primer_dia
    created = []
    while curr <= ultimo_dia:
        # solo considerar días laborales (lunes-viernes)
        if curr.weekday() < 5:
            # buscar último evento 'out' de ese día
            ev = AsistenciaEvento.query.filter(
                AsistenciaEvento.empleado_id == empleado.id,
                func.date(AsistenciaEvento.ts) == curr,
                AsistenciaEvento.tipo == 'out'
            ).order_by(AsistenciaEvento.ts.desc()).first()

            if ev:
                # calcular diferencia en horas
                fecha_fin_jornada = datetime(curr.year, curr.month, curr.day, fin_jornada_hour, 0, 0)
                if ev.ts > fecha_fin_jornada:
                    delta = ev.ts - fecha_fin_jornada
                    horas_decimal = round(delta.total_seconds() / 3600.0, 4)

                    # calcular monto
                    salario_hora = float(empleado.salario_base) / float(dias_habiles) / float(horas_jornada)
                    multiplicador = 1.5
                    monto = round(horas_decimal * salario_hora * multiplicador)

                    # crear o actualizar registro HorasExtra
                    he = HorasExtra.query.filter_by(empleado_id=empleado.id, fecha=curr).first()
                    if he:
                        he.horas = horas_decimal
                        he.monto_calculado = monto
                        he.origen = 'asistencia'
                    else:
                        he = HorasExtra(
                            empleado_id=empleado.id,
                            fecha=curr,
                            horas=horas_decimal,
                            monto_calculado=monto,
                            origen='asistencia',
                            estado='PENDIENTE'
                        )
                        db.session.add(he)
                        created.append(he)
        curr = curr + timedelta(days=1)

    if created:
        db.session.commit()
    return True


@rrhh_bp.route('/api/ingresos-extras/employees', methods=['GET'])
@login_required
def api_ingresos_extras_employees():
    """Devuelve lista de empleados con resumen de horas extras (mes actual por defecto)"""
    periodo = request.args.get('periodo', date.today().strftime('%Y-%m'))  # YYYY-MM
    año, mes = map(int, periodo.split('-'))

    empleados = Empleado.query.order_by(Empleado.apellido, Empleado.nombre).all()
    items = []
    for e in empleados:
        total_horas = db.session.query(func.coalesce(func.sum(HorasExtra.horas), 0)).filter(
            HorasExtra.empleado_id == e.id,
            func.extract('year', HorasExtra.fecha) == año,
            func.extract('month', HorasExtra.fecha) == mes,
            HorasExtra.estado == 'PENDIENTE'
        ).scalar() or 0

        pendientes = db.session.query(func.count(HorasExtra.id)).filter(
            HorasExtra.empleado_id == e.id,
            func.extract('year', HorasExtra.fecha) == año,
            func.extract('month', HorasExtra.fecha) == mes,
            HorasExtra.estado == 'PENDIENTE'
        ).scalar() or 0

        items.append({'id': e.id, 'nombre': e.nombre_completo, 'pendientes': int(pendientes), 'total_horas': float(total_horas)})

    return jsonify({'items': items, 'periodo': periodo})


@rrhh_bp.route('/api/empleados/<int:empleado_id>/horas-extras', methods=['GET'])
@login_required
def api_empleado_horas_extras(empleado_id):
    """Devuelve horas extras de un empleado en un mes; si no hay registros intenta detectar desde eventos de asistencia."""
    periodo = request.args.get('periodo', date.today().strftime('%Y-%m'))
    año, mes = map(int, periodo.split('-'))

    empleado = Empleado.query.get_or_404(empleado_id)

    # intentar detectar/crear registros desde eventos de asistencia
    _compute_horas_extra_for_month(empleado, año, mes)

    horas = HorasExtra.query.filter(
        HorasExtra.empleado_id == empleado_id,
        func.extract('year', HorasExtra.fecha) == año,
        func.extract('month', HorasExtra.fecha) == mes
    ).order_by(HorasExtra.fecha.desc()).all()

    items = []
    for h in horas:
        items.append({
            'id': h.id,
            'fecha': h.fecha.strftime('%d/%m/%Y'),
            'horas': float(h.horas),
            'monto': float(h.monto_calculado),
            'estado': h.estado
        })

    return jsonify({'items': items, 'empleado': empleado.nombre_completo, 'periodo': periodo})


@rrhh_bp.route('/horas-extra/<int:hora_id>/approve', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def aprobar_hora_extra(hora_id):
    he = HorasExtra.query.get_or_404(hora_id)
    try:
        he.estado = 'APROBADO'
        he.aprobado_por = current_user.id
        he.fecha_aprobacion = datetime.utcnow()
        db.session.commit()
        return jsonify({'message': 'Hora extra aprobada'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@rrhh_bp.route('/horas-extra/<int:hora_id>/reject', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def rechazar_hora_extra(hora_id):
    he = HorasExtra.query.get_or_404(hora_id)
    try:
        he.estado = 'RECHAZADO'
        he.aprobado_por = current_user.id
        he.fecha_aprobacion = datetime.utcnow()
        db.session.commit()
        return jsonify({'message': 'Hora extra rechazada'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@rrhh_bp.route('/liquidaciones/preview/<periodo>', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def preview_liquidacion(periodo):
    """Pre-visualización de liquidación antes de generar"""
    try:
        año, mes = map(int, periodo.split('-'))
        empleados = Empleado.query.filter_by(estado=EstadoEmpleadoEnum.ACTIVO).all()
        
        preview_data = []
        totales = {
            'salarios': Decimal('0'),
            'bonificaciones': Decimal('0'),
            'ingresos': Decimal('0'),
            'descuentos': Decimal('0'),
            'anticipos': Decimal('0'),
            'ips': Decimal('0'),
            'neto': Decimal('0')
        }
        
        for empleado in empleados:
            # Calcular días presentes
            dias_presentes = db.session.query(func.count(Asistencia.id)).filter(
                Asistencia.empleado_id == empleado.id,
                func.extract('month', Asistencia.fecha) == mes,
                func.extract('year', Asistencia.fecha) == año,
                Asistencia.presente == True
            ).scalar() or 0
            
            salario_proporcional = (empleado.salario_base / 30) * dias_presentes
            
            # Ingresos extras
            ingresos = db.session.query(func.coalesce(func.sum(IngresoExtra.monto), 0)).filter(
                IngresoExtra.empleado_id == empleado.id,
                IngresoExtra.mes == mes,
                IngresoExtra.año == año,
                IngresoExtra.estado == 'APROBADO',
                IngresoExtra.aplicado == False
            ).scalar() or Decimal('0')
            
            # Bonificación
            bonificacion = calcular_bonificacion_familiar(empleado.id, date(año, mes, 1))
            
            # Descuentos
            descuentos = db.session.query(func.sum(Descuento.monto)).filter(
                Descuento.empleado_id == empleado.id,
                Descuento.mes == mes,
                Descuento.año == año
            ).scalar() or Decimal('0')
            
            # Anticipos
            anticipos = db.session.query(func.sum(Anticipo.monto)).filter(
                Anticipo.empleado_id == empleado.id,
                func.extract('month', Anticipo.fecha_aprobacion) == mes,
                func.extract('year', Anticipo.fecha_aprobacion) == año,
                Anticipo.aprobado == True,
                Anticipo.aplicado == False
            ).scalar() or Decimal('0')
            
            descuentos_total = descuentos + anticipos
            ips = (salario_proporcional + ingresos + bonificacion) * Decimal('0.09625')
            neto = salario_proporcional + ingresos + bonificacion - descuentos_total - ips
            
            preview_data.append({
                'codigo': empleado.codigo,
                'nombre': empleado.nombre_completo,
                'dias': dias_presentes,
                'salario': float(salario_proporcional),
                'bonificacion': float(bonificacion),
                'ingresos': float(ingresos),
                'descuentos': float(descuentos),
                'anticipos': float(anticipos),
                'ips': float(ips),
                'neto': float(neto)
            })
            
            totales['salarios'] += salario_proporcional
            totales['bonificaciones'] += bonificacion
            totales['ingresos'] += ingresos
            totales['descuentos'] += descuentos
            totales['anticipos'] += anticipos
            totales['ips'] += ips
            totales['neto'] += neto
        
        return jsonify({
            'periodo': periodo,
            'empleados': preview_data,
            'totales': {k: float(v) for k, v in totales.items()},
            'cantidad_empleados': len(preview_data)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@rrhh_bp.route('/anticipos/pendientes', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def anticipos_pendientes():
    """Reporte de anticipos pendientes de aplicar"""
    anticipos = db.session.query(
        Anticipo, Empleado
    ).join(
        Empleado, Anticipo.empleado_id == Empleado.id
    ).filter(
        Anticipo.aprobado == True,
        Anticipo.aplicado == False
    ).order_by(Anticipo.fecha_aprobacion.desc()).all()
    
    datos = []
    total = Decimal('0')
    for anticipo, empleado in anticipos:
        periodo_descuento = anticipo.fecha_aprobacion.strftime('%Y-%m')
        
        # Verificar si ya se generó liquidación para ese período
        liquidacion = Liquidacion.query.filter_by(
            empleado_id=empleado.id,
            periodo=periodo_descuento
        ).first()
        
        datos.append({
            'id': anticipo.id,
            'empleado_codigo': empleado.codigo,
            'empleado_nombre': empleado.nombre_completo,
            'monto': float(anticipo.monto),
            'fecha_aprobacion': anticipo.fecha_aprobacion.strftime('%d/%m/%Y'),
            'periodo_a_descontar': periodo_descuento,
            'estado': 'Ya liquidado (no descontado)' if liquidacion else 'Pendiente de liquidación'
        })
        total += anticipo.monto
    
    return jsonify({
        'anticipos': datos,
        'total': float(total),
        'cantidad': len(datos)
    })

@rrhh_bp.route('/metricas/asistencias', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def metricas_asistencias():
    """Dashboard de métricas de asistencias"""
    mes = request.args.get('mes', date.today().month, type=int)
    año = request.args.get('year', date.today().year, type=int)
    
    # Calcular días hábiles del mes
    import calendar
    primer_dia = date(año, mes, 1)
    ultimo_dia = date(año, mes, calendar.monthrange(año, mes)[1])
    
    dias_habiles = 0
    fecha_actual = primer_dia
    while fecha_actual <= ultimo_dia:
        if fecha_actual.weekday() < 5:  # Lunes a viernes
            dias_habiles += 1
        fecha_actual += timedelta(days=1)
    
    # Métricas por empleado
    empleados = Empleado.query.filter_by(estado=EstadoEmpleadoEnum.ACTIVO).all()
    metricas = []
    
    for empleado in empleados:
        dias_presentes = db.session.query(func.count(Asistencia.id)).filter(
            Asistencia.empleado_id == empleado.id,
            func.extract('month', Asistencia.fecha) == mes,
            func.extract('year', Asistencia.fecha) == año,
            Asistencia.presente == True
        ).scalar() or 0
        
        ausencias = db.session.query(func.count(Asistencia.id)).filter(
            Asistencia.empleado_id == empleado.id,
            func.extract('month', Asistencia.fecha) == mes,
            func.extract('year', Asistencia.fecha) == año,
            Asistencia.presente == False
        ).scalar() or 0
        
        ausencias_justificadas = db.session.query(func.count(Asistencia.id)).filter(
            Asistencia.empleado_id == empleado.id,
            func.extract('month', Asistencia.fecha) == mes,
            func.extract('year', Asistencia.fecha) == año,
            Asistencia.presente == False,
            Asistencia.justificacion_estado == 'JUSTIFICADO'
        ).scalar() or 0
        
        ausencias_injustificadas = db.session.query(func.count(Asistencia.id)).filter(
            Asistencia.empleado_id == empleado.id,
            func.extract('month', Asistencia.fecha) == mes,
            func.extract('year', Asistencia.fecha) == año,
            Asistencia.presente == False,
            Asistencia.justificacion_estado == 'INJUSTIFICADO'
        ).scalar() or 0
        
        tasa_asistencia = (dias_presentes / dias_habiles * 100) if dias_habiles > 0 else 0
        
        metricas.append({
            'empleado_id': empleado.id,
            'codigo': empleado.codigo,
            'nombre': empleado.nombre_completo,
            'dias_presentes': dias_presentes,
            'ausencias_totales': ausencias,
            'ausencias_justificadas': ausencias_justificadas,
            'ausencias_injustificadas': ausencias_injustificadas,
            'tasa_asistencia': round(tasa_asistencia, 2)
        })
    
    # Ordenar por ausencias injustificadas (mayor a menor)
    metricas.sort(key=lambda x: x['ausencias_injustificadas'], reverse=True)
    
    # Resumen general
    total_presentes = sum(m['dias_presentes'] for m in metricas)
    total_ausencias = sum(m['ausencias_totales'] for m in metricas)
    total_justificadas = sum(m['ausencias_justificadas'] for m in metricas)
    total_injustificadas = sum(m['ausencias_injustificadas'] for m in metricas)
    
    return jsonify({
        'periodo': f'{año}-{mes:02d}',
        'dias_habiles': dias_habiles,
        'empleados': metricas,
        'resumen': {
            'total_empleados': len(metricas),
            'total_presentes': total_presentes,
            'total_ausencias': total_ausencias,
            'total_justificadas': total_justificadas,
            'total_injustificadas': total_injustificadas,
            'tasa_general': round((total_presentes / (len(metricas) * dias_habiles) * 100) if dias_habiles > 0 and len(metricas) > 0 else 0, 2)
        }
    })

@rrhh_bp.route('/liquidaciones/generar', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def generar_liquidacion():
    """Generar liquidaciones para un período"""
    if request.method == 'POST':
        try:
            periodo = request.form.get('periodo')  # YYYY-MM
            año, mes = map(int, periodo.split('-'))
            
            # Obtener todos los empleados activos
            empleados = Empleado.query.filter_by(estado=EstadoEmpleadoEnum.ACTIVO).all()
            
            contador = 0
            for empleado in empleados:
                # Evitar duplicados
                liquidacion_existente = Liquidacion.query.filter_by(
                    empleado_id=empleado.id,
                    periodo=periodo
                ).first()
                
                if liquidacion_existente:
                    continue
                
                # ========================================
                # CALCULAR DÍAS TRABAJADOS DESDE ASISTENCIAS
                # ========================================
                dias_presentes = db.session.query(func.count(Asistencia.id)).filter(
                    Asistencia.empleado_id == empleado.id,
                    func.extract('month', Asistencia.fecha) == mes,
                    func.extract('year', Asistencia.fecha) == año,
                    Asistencia.presente == True
                ).scalar() or 0
                
                # Calcular días hábiles teóricos del mes (lunes a viernes)
                import calendar
                primer_dia = date(año, mes, 1)
                ultimo_dia = date(año, mes, calendar.monthrange(año, mes)[1])
                
                dias_habiles_teoricos = 0
                fecha_actual = primer_dia
                while fecha_actual <= ultimo_dia:
                    if fecha_actual.weekday() < 5:  # Lunes a viernes
                        dias_habiles_teoricos += 1
                    fecha_actual += timedelta(days=1)
                
                dias_ausentes = dias_habiles_teoricos - dias_presentes
                
                # ========================================
                # VALIDACIÓN: Días presentes no puede superar días hábiles
                # ========================================
                if dias_presentes > dias_habiles_teoricos:
                    print(f"⚠️ ALERTA: Empleado {empleado.codigo} tiene {dias_presentes} días presentes pero solo hay {dias_habiles_teoricos} días hábiles")
                    flash(f'Advertencia: {empleado.nombre_completo} tiene inconsistencia en asistencias ({dias_presentes} > {dias_habiles_teoricos})', 'warning')
                
                # ========================================
                # CALCULAR SALARIO PROPORCIONAL
                # ========================================
                print(f"\n{'='*60}")
                print(f"📊 LIQUIDACIÓN: {empleado.codigo} - {empleado.nombre_completo}")
                print(f"{'='*60}")
                salario_diario = empleado.salario_base / Decimal(30)
                salario_base_ajustado = salario_diario * Decimal(str(dias_presentes))
                print(f"💰 Salario base: ₲{empleado.salario_base:,.2f}")
                print(f"📅 Días presentes: {dias_presentes}/{dias_habiles_teoricos}")
                print(f"💵 Salario proporcional: ₲{salario_base_ajustado:,.2f}")
                
                # Calcular ingresos extras: incluir solo ingresos manuales APROBADOS y horas extra APROBADAS
                ingresos_extras = db.session.query(func.coalesce(func.sum(IngresoExtra.monto), 0)).filter(
                    IngresoExtra.empleado_id == empleado.id,
                    IngresoExtra.mes == mes,
                    IngresoExtra.año == año,
                    IngresoExtra.estado == 'APROBADO',
                    IngresoExtra.aplicado == False
                ).scalar() or Decimal('0')

                # Sumar horas extra aprobadas y no aplicadas
                horas_extra_total = db.session.query(func.coalesce(func.sum(HorasExtra.monto_calculado), 0)).filter(
                    HorasExtra.empleado_id == empleado.id,
                    func.extract('year', HorasExtra.fecha) == año,
                    func.extract('month', HorasExtra.fecha) == mes,
                    HorasExtra.estado == 'APROBADO',
                    HorasExtra.aplicado == False
                ).scalar() or Decimal('0')

                # Asegurar Decimal y sumar
                ingresos_extras = Decimal(str(ingresos_extras)) + Decimal(str(horas_extra_total))
                print(f"➕ Ingresos extras: ₲{ingresos_extras:,.2f}")
                
                # Calcular descuentos manuales
                descuentos = db.session.query(func.sum(Descuento.monto)).filter(
                    Descuento.empleado_id == empleado.id,
                    Descuento.mes == mes,
                    Descuento.año == año
                ).scalar() or Decimal('0')
                print(f"➖ Descuentos manuales: ₲{descuentos:,.2f}")
                
                # 🆕 CRÍTICO: Calcular anticipos aprobados del mes que no se han aplicado
                anticipos_mes = db.session.query(func.sum(Anticipo.monto)).filter(
                    Anticipo.empleado_id == empleado.id,
                    func.extract('month', Anticipo.fecha_aprobacion) == mes,
                    func.extract('year', Anticipo.fecha_aprobacion) == año,
                    Anticipo.aprobado == True,
                    Anticipo.aplicado == False
                ).scalar() or Decimal('0')
                print(f"➖ Anticipos del mes: ₲{anticipos_mes:,.2f}")
                
                # Sumar descuentos totales
                descuentos_totales = descuentos + anticipos_mes
                print(f"➖ TOTAL DESCUENTOS: ₲{descuentos_totales:,.2f}")
                
                # Calcular bonificación familiar
                fecha_liquidacion = date(año, mes, 1)
                bonificacion_familiar = calcular_bonificacion_familiar(empleado.id, fecha_liquidacion)
                print(f"👨‍👩‍👧 Bonificación familiar: ₲{bonificacion_familiar:,.2f}")
                
                # Calcular aporte IPS (9.625% sobre salario ajustado + ingresos extras + bonificación)
                aporte_ips = (salario_base_ajustado + ingresos_extras + bonificacion_familiar) * Decimal('0.09625')
                print(f"🏦 Aporte IPS (9.625%): ₲{aporte_ips:,.2f}")
                
                # Calcular salario neto (USAR descuentos_totales)
                salario_neto = salario_base_ajustado + ingresos_extras + bonificacion_familiar - descuentos_totales - aporte_ips
                print(f"{'='*60}")
                print(f"💵 SALARIO NETO: ₲{salario_neto:,.2f}")
                print(f"{'='*60}\n")
                
                # Crear liquidación (USAR descuentos_totales)
                liquidacion = Liquidacion(
                    empleado_id=empleado.id,
                    periodo=periodo,
                    salario_base=salario_base_ajustado,  # Salario ajustado a días trabajados
                    ingresos_extras=ingresos_extras,
                    bonificacion_familiar=bonificacion_familiar,
                    descuentos=descuentos_totales,  # ← CRÍTICO: Incluye anticipos
                    aporte_ips=aporte_ips,
                    salario_neto=salario_neto,
                    dias_trabajados=dias_presentes
                )
                
                db.session.add(liquidacion)

                # Marcar los IngresoExtra aplicados (APROBADOS y no aplicados)
                ingresos_a_aplicar = IngresoExtra.query.filter_by(
                    empleado_id=empleado.id,
                    mes=mes,
                    año=año,
                    estado='APROBADO',
                    aplicado=False
                ).all()
                for ie in ingresos_a_aplicar:
                    ie.aplicado = True
                    ie.fecha_aplicacion = datetime.utcnow()

                # Marcar HorasExtra aplicadas (APROBADO y no aplicadas)
                horas_a_aplicar = HorasExtra.query.filter(
                    HorasExtra.empleado_id == empleado.id,
                    func.extract('year', HorasExtra.fecha) == año,
                    func.extract('month', HorasExtra.fecha) == mes,
                    HorasExtra.estado == 'APROBADO',
                    HorasExtra.aplicado == False
                ).all()
                for he in horas_a_aplicar:
                    he.aplicado = True
                    he.fecha_aplicacion = datetime.utcnow()
                
                # 🆕 CRÍTICO: Marcar anticipos como aplicados
                anticipos_a_aplicar = Anticipo.query.filter(
                    Anticipo.empleado_id == empleado.id,
                    func.extract('month', Anticipo.fecha_aprobacion) == mes,
                    func.extract('year', Anticipo.fecha_aprobacion) == año,
                    Anticipo.aprobado == True,
                    Anticipo.aplicado == False
                ).all()
                for anticipo in anticipos_a_aplicar:
                    anticipo.aplicado = True
                    anticipo.fecha_aplicacion = date(año, mes, 1)
                    print(f"✅ Anticipo ID {anticipo.id} marcado como aplicado")

                contador += 1
            
            db.session.commit()
            
            registrar_bitacora(
                current_user, 'liquidaciones', 'CREATE', 'liquidaciones',
                detalle=f'Generadas {contador} liquidaciones para período {periodo}'
            )
            
            flash(f'{contador} liquidaciones generadas exitosamente (cálculo basado en asistencias)', 'success')
            return redirect(url_for('rrhh.listar_liquidaciones', periodo=periodo))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al generar liquidaciones: {str(e)}', 'danger')
    
    return render_template('rrhh/generar_liquidacion.html')

@rrhh_bp.route('/liquidaciones/<int:liquidacion_id>/descargar-pdf', methods=['GET'])
@login_required
def descargar_recibo_pdf(liquidacion_id):
    """Descargar recibo de salario en PDF"""
    liquidacion = Liquidacion.query.get_or_404(liquidacion_id)
    empleado = liquidacion.empleado
    
    registrar_bitacora(
        current_user, 'liquidaciones', 'VIEW', 'liquidaciones',
        liquidacion_id, 'Descarga de recibo PDF'
    )
    
    # Generar PDF
    pdf_buffer = ReportUtils.generar_recibo_salario(empleado, liquidacion)
    
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'recibo_{empleado.codigo}_{liquidacion.periodo}.pdf'
    )

@rrhh_bp.route('/liquidaciones/planilla-mensual/<periodo>/pdf', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def descargar_planilla_mensual(periodo):
    """Descargar planilla mensual consolidada"""
    liquidaciones = Liquidacion.query.filter_by(periodo=periodo).all()
    empleados_liquidaciones = [(l.empleado, l) for l in liquidaciones]
    
    if not empleados_liquidaciones:
        flash('No hay liquidaciones para ese período', 'warning')
        return redirect(url_for('rrhh.listar_liquidaciones', periodo=periodo))
    
    registrar_bitacora(
        current_user, 'liquidaciones', 'VIEW', 'liquidaciones',
        detalle=f'Descarga de planilla mensual {periodo}'
    )
    
    pdf_buffer = ReportUtils.generar_planilla_mensual(empleados_liquidaciones, periodo)
    
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'planilla_mensual_{periodo}.pdf'
    )

# ==================== VACACIONES ====================

def calcular_dias_vacaciones_por_antiguedad(empleado, año=None):
    """
    Calcula días de vacaciones según antigüedad del empleado.
    Código Laboral Paraguayo:
    - 1-5 años: 12 días
    - 5-10 años: 18 días
    - 10+ años: 30 días
    
    Args:
        empleado: Objeto Empleado
        año: Año para el cual calcular (por defecto año actual)
    
    Returns:
        int: Días de vacaciones correspondientes
    """
    if not año:
        año = date.today().year
    
    if not empleado.fecha_ingreso:
        return 12  # Default mínimo
    
    # Calcular antigüedad al 31 de diciembre del año solicitado
    fecha_calculo = date(año, 12, 31)
    if fecha_calculo > date.today():
        fecha_calculo = date.today()
    
    años_servicio = (fecha_calculo - empleado.fecha_ingreso).days / 365.25
    
    if años_servicio < 5:
        return 12
    elif años_servicio < 10:
        return 18
    else:
        return 30


def calcular_saldo_vacaciones_actual(empleado):
    """
    Calcula el saldo actual de vacaciones del empleado dinámicamente.
    
    - Calcula días ganados por antigüedad cada año
    - Resta días tomados (aprobados)
    - Acumula hasta 2 años atrás
    - Identifica días por vencer
    
    Args:
        empleado: Objeto Empleado
    
    Returns:
        dict: {
            'disponibles_total': int,
            'tomados_total': int,
            'pendientes': int,
            'por_vencer': int,
            'fecha_vencimiento': date,
            'alerta': bool,
            'desglose_años': list
        }
    """
    if not empleado.fecha_ingreso:
        return {
            'disponibles_total': 0,
            'tomados_total': 0,
            'pendientes': 0,
            'por_vencer': 0,
            'fecha_vencimiento': None,
            'alerta': False,
            'desglose_años': []
        }
    
    año_actual = date.today().year
    años_a_considerar = [año_actual - 2, año_actual - 1, año_actual]  # 3 años (incluye el que vence)
    
    disponibles_total = 0
    tomados_total = 0
    desglose_años = []
    por_vencer = 0
    fecha_vencimiento = None
    
    for año in años_a_considerar:
        # Solo considerar años después del ingreso
        if año < empleado.fecha_ingreso.year:
            continue
        
        # Días ganados ese año según antigüedad
        dias_ganados = calcular_dias_vacaciones_por_antiguedad(empleado, año)
        
        # Días tomados ese año (vacaciones aprobadas)
        vacaciones_año = Vacacion.query.filter_by(
            empleado_id=empleado.id,
            año=año,
            estado=EstadoVacacionEnum.APROBADA
        ).all()
        
        dias_tomados_año = sum([
            (v.fecha_fin_solicitud - v.fecha_inicio_solicitud).days + 1
            for v in vacaciones_año
            if v.fecha_inicio_solicitud and v.fecha_fin_solicitud
        ])
        
        saldo_año = dias_ganados - dias_tomados_año
        
        # Identificar días por vencer (más de 2 años)
        if año == año_actual - 2 and saldo_año > 0:
            por_vencer = saldo_año
            fecha_vencimiento = date(año_actual, 12, 31)
        
        desglose_años.append({
            'año': año,
            'ganados': dias_ganados,
            'tomados': dias_tomados_año,
            'saldo': saldo_año,
            'vence': año == año_actual - 2
        })
        
        disponibles_total += dias_ganados
        tomados_total += dias_tomados_año
    
    pendientes = disponibles_total - tomados_total
    alerta = por_vencer > 0
    
    return {
        'disponibles_total': disponibles_total,
        'tomados_total': tomados_total,
        'pendientes': max(0, pendientes),
        'por_vencer': por_vencer,
        'fecha_vencimiento': fecha_vencimiento,
        'alerta': alerta,
        'desglose_años': desglose_años
    }


def generar_vacaciones_anuales(año=None, empleado_id=None):
    """
    Genera registros de vacaciones para un año específico.
    - Calcula días según antigüedad de cada empleado
    - Acumula saldos pendientes del año anterior
    - Crea registro solo si no existe
    
    Args:
        año: Año para generar (por defecto año actual)
        empleado_id: ID específico o None para todos los activos
    
    Returns:
        dict: Resumen de generación (creados, actualizados, errores)
    """
    if not año:
        año = date.today().year
    
    # Obtener empleados
    if empleado_id:
        empleados = [Empleado.query.get(empleado_id)]
    else:
        empleados = Empleado.query.filter_by(estado=EstadoEmpleadoEnum.ACTIVO).all()
    
    resultado = {
        'creados': 0,
        'actualizados': 0,
        'ya_existentes': 0,
        'errores': []
    }
    
    for empleado in empleados:
        try:
            # Verificar si ya existe registro para ese año
            vacacion_existente = Vacacion.query.filter_by(
                empleado_id=empleado.id,
                año=año
            ).first()
            
            if vacacion_existente:
                # Actualizar días disponibles según antigüedad actual
                dias_por_antiguedad = calcular_dias_vacaciones_por_antiguedad(empleado, año)
                
                # Solo actualizar si cambió la antigüedad
                if vacacion_existente.dias_disponibles != dias_por_antiguedad:
                    diferencia = dias_por_antiguedad - (vacacion_existente.dias_disponibles - vacacion_existente.dias_tomados)
                    vacacion_existente.dias_disponibles = dias_por_antiguedad
                    vacacion_existente.dias_pendientes = max(0, dias_por_antiguedad - vacacion_existente.dias_tomados)
                    db.session.commit()
                    resultado['actualizados'] += 1
                else:
                    resultado['ya_existentes'] += 1
                continue
            
            # Calcular días según antigüedad
            dias_disponibles = calcular_dias_vacaciones_por_antiguedad(empleado, año)
            
            # Buscar saldo pendiente del año anterior
            año_anterior = año - 1
            vacacion_anterior = Vacacion.query.filter_by(
                empleado_id=empleado.id,
                año=año_anterior
            ).first()
            
            saldo_anterior = 0
            if vacacion_anterior and vacacion_anterior.dias_pendientes > 0:
                # Acumular hasta 2 años de vacaciones no gozadas (Paraguay permite acumulación)
                saldo_anterior = min(vacacion_anterior.dias_pendientes, dias_disponibles * 2)
            
            # Crear nuevo registro
            nueva_vacacion = Vacacion(
                empleado_id=empleado.id,
                año=año,
                dias_disponibles=dias_disponibles + saldo_anterior,
                dias_tomados=0,
                dias_pendientes=dias_disponibles + saldo_anterior,
                estado=EstadoVacacionEnum.PENDIENTE
            )
            
            db.session.add(nueva_vacacion)
            db.session.commit()
            resultado['creados'] += 1
            
        except Exception as e:
            resultado['errores'].append(f"Empleado {empleado.codigo}: {str(e)}")
            db.session.rollback()
    
    return resultado


@rrhh_bp.route('/vacaciones', methods=['GET'])
@login_required
def listar_vacaciones():
    """Lista vacaciones"""
    registrar_bitacora(current_user, 'vacaciones', 'VIEW', 'vacaciones')
    
    pagina = request.args.get('page', 1, type=int)
    # Normalizar el filtro de año a int (request.args devuelve str si viene por querystring)
    año_filtro_raw = request.args.get('año', date.today().year)
    try:
        año_filtro = int(año_filtro_raw)
    except (TypeError, ValueError):
        año_filtro = date.today().year

    vacaciones = Vacacion.query.filter_by(año=año_filtro).paginate(page=pagina, per_page=10)
    
    return render_template('rrhh/vacaciones.html', vacaciones=vacaciones, año=año_filtro)


@rrhh_bp.route('/vacaciones/empleado/<int:empleado_id>', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def ver_historial_vacaciones(empleado_id):
    """Ver historial de vacaciones de un empleado con cálculo dinámico y alertas"""
    registrar_bitacora(current_user, 'vacaciones', 'VIEW', 'vacaciones')

    empleado = Empleado.query.get_or_404(empleado_id)
    
    # Calcular saldo actual dinámicamente
    saldo = calcular_saldo_vacaciones_actual(empleado)
    
    # Traer todas las vacaciones (historial) del empleado
    vacaciones = Vacacion.query.filter_by(empleado_id=empleado_id).order_by(Vacacion.año.desc()).all()
    # Traer solicitudes pendientes
    pendientes = Vacacion.query.filter_by(empleado_id=empleado_id, estado=EstadoVacacionEnum.PENDIENTE).all()
    
    # Calcular días por antigüedad para mostrar
    empleado.dias_vacaciones_actuales = calcular_dias_vacaciones_por_antiguedad(empleado)

    return render_template('rrhh/vacacion_detalle.html', 
                         empleado=empleado, 
                         vacaciones=vacaciones, 
                         pendientes=pendientes,
                         saldo=saldo)


@rrhh_bp.route('/vacaciones/solicitud_imprimir', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def imprimir_solicitud_vacacion():
    """Generar PDF de solicitud de vacaciones para imprimir y firmar"""
    empleados = Empleado.query.filter_by(estado=EstadoEmpleadoEnum.ACTIVO).all()

    if request.method == 'POST':
        try:
            empleado_id = int(request.form.get('empleado_id'))
            empleado = Empleado.query.get_or_404(empleado_id)

            # Priorizar si viene vacacion_id (imprimir solicitud existente)
            vacacion_id = request.form.get('vacacion_id')
            if vacacion_id:
                vacacion = Vacacion.query.get(int(vacacion_id))
                fecha_inicio = vacacion.fecha_inicio_solicitud
                fecha_fin = vacacion.fecha_fin_solicitud
                dias = (fecha_fin - fecha_inicio).days + 1 if fecha_inicio and fecha_fin else ''
            else:
                # Tomar datos manuales
                fecha_inicio = request.form.get('fecha_inicio') and datetime.strptime(request.form.get('fecha_inicio'), '%Y-%m-%d').date()
                fecha_fin = request.form.get('fecha_fin') and datetime.strptime(request.form.get('fecha_fin'), '%Y-%m-%d').date()
                dias = (fecha_fin - fecha_inicio).days + 1 if fecha_inicio and fecha_fin else ''

            # Generar PDF
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)
            width, height = A4

            # Encabezado
            c.setFont('Helvetica-Bold', 14)
            c.drawString(50, height - 50, 'Solicitud de Vacaciones')
            c.setFont('Helvetica', 10)
            c.drawString(50, height - 80, f'Fecha de solicitud: {datetime.now().strftime("%d/%m/%Y")}')

            # Datos del empleado
            c.setFont('Helvetica-Bold', 11)
            c.drawString(50, height - 120, 'Datos del funcionario:')
            c.setFont('Helvetica', 10)
            c.drawString(60, height - 140, f'Nombre: {empleado.nombre}')
            c.drawString(300, height - 140, f'Apellido: {empleado.apellido}')
            c.drawString(60, height - 160, f'Código: {empleado.codigo}')
            c.drawString(60, height - 180, f'Días a solicitar: {dias}')
            c.drawString(60, height - 200, f'Periodo solicitado: {fecha_inicio.strftime("%d/%m/%Y") if fecha_inicio else "-"} - {fecha_fin.strftime("%d/%m/%Y") if fecha_fin else "-"}')

            # Espacio para firmas
            y_sig = height - 260
            line_y_gap = 60
            c.line(60, y_sig, 260, y_sig)
            c.drawString(60, y_sig - 15, 'Firma del funcionario')

            c.line(300, y_sig, 500, y_sig)
            c.drawString(300, y_sig - 15, 'Firma del jefe')

            c.line(60, y_sig - line_y_gap, 260, y_sig - line_y_gap)
            c.drawString(60, y_sig - line_y_gap - 15, 'Firma RRHH')

            # Pie
            c.setFont('Helvetica-Oblique', 9)
            c.drawString(50, 40, 'Este documento es la solicitud formal del funcionario para gestionar sus vacaciones ante RRHH.')

            c.showPage()
            c.save()
            buffer.seek(0)

            filename = f'solicitud_vacaciones_{empleado.codigo}_{datetime.now().strftime("%Y%m%d")}.pdf'
            return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

        except Exception as e:
            flash(f'Error generando PDF: {e}', 'danger')

    # GET
    return render_template('rrhh/vacacion_imprimir.html', empleados=empleados)


@rrhh_bp.route('/vacaciones/solicitud_imprimir_modelo', methods=['GET'])
@login_required
def imprimir_solicitud_vacacion_modelo():
    """Generar y descargar el PDF modelo en blanco para que el funcionario lo complete manualmente."""
    try:
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        c.setFont('Helvetica-Bold', 16)
        c.drawString(50, height - 50, 'Solicitud de Vacaciones (Modelo)')
        c.setFont('Helvetica', 10)
        c.drawString(50, height - 80, f'Fecha: {datetime.now().strftime("%d/%m/%Y")}')

        # Campos en blanco para completar
        y = height - 120
        field_gap = 24
        c.drawString(50, y, 'Nombre:')
        c.line(120, y - 3, 500, y - 3)
        y -= field_gap

        c.drawString(50, y, 'Apellido:')
        c.line(120, y - 3, 500, y - 3)
        y -= field_gap

        c.drawString(50, y, 'Código:')
        c.line(120, y - 3, 300, y - 3)
        y -= field_gap

        c.drawString(50, y, 'Días a solicitar:')
        c.line(150, y - 3, 220, y - 3)
        y -= field_gap

        c.drawString(50, y, 'Periodo solicitado (desde - hasta):')
        c.line(240, y - 3, 500, y - 3)
        y -= (field_gap + 10)

        # Espacio para firmas
        sig_y = y - 40
        c.line(60, sig_y, 260, sig_y)
        c.drawString(60, sig_y - 15, 'Firma del funcionario')

        c.line(300, sig_y, 500, sig_y)
        c.drawString(300, sig_y - 15, 'Firma del jefe')

        c.line(60, sig_y - 80, 260, sig_y - 80)
        c.drawString(60, sig_y - 95, 'Firma RRHH')

        c.setFont('Helvetica-Oblique', 9)
        c.drawString(50, 40, 'Imprimir y entregar este formulario completo a RRHH para la gestión de la solicitud.')

        c.showPage()
        c.save()
        buffer.seek(0)
        filename = f'modelo_solicitud_vacaciones_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

    except Exception as e:
        flash(f'Error generando PDF modelo: {e}', 'danger')
        return redirect(url_for('rrhh.listar_vacaciones'))


@rrhh_bp.route('/vacaciones/solicitud_imprimir_directa/<int:empleado_id>', methods=['GET'])
@login_required
def imprimir_solicitud_vacacion_directa(empleado_id):
    """Generar PDF prellenado para un empleado y descargarlo directamente (opcional vacacion_id en querystring)."""
    try:
        empleado = Empleado.query.get_or_404(empleado_id)
        vacacion_id = request.args.get('vacacion_id')
        fecha_inicio = fecha_fin = None
        dias = ''
        if vacacion_id:
            vac = Vacacion.query.get(int(vacacion_id))
            if vac:
                fecha_inicio = vac.fecha_inicio_solicitud
                fecha_fin = vac.fecha_fin_solicitud
                if fecha_inicio and fecha_fin:
                    dias = (fecha_fin - fecha_inicio).days + 1

        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        # Encabezado simple
        c.setFont('Helvetica-Bold', 14)
        c.drawString(50, height - 50, 'Solicitud de Vacaciones')
        c.setFont('Helvetica', 10)
        c.drawString(50, height - 80, f'Fecha de solicitud: {datetime.now().strftime("%d/%m/%Y")}')

        # Datos del empleado prellenados
        c.setFont('Helvetica-Bold', 11)
        c.drawString(50, height - 120, 'Datos del funcionario:')
        c.setFont('Helvetica', 10)
        c.drawString(60, height - 140, f'Nombre: {empleado.nombre}')
        c.drawString(300, height - 140, f'Apellido: {empleado.apellido}')
        c.drawString(60, height - 160, f'Código: {empleado.codigo}')
        c.drawString(60, height - 180, f'Días a solicitar: {dias}')
        c.drawString(60, height - 200, f'Periodo solicitado: {fecha_inicio.strftime("%d/%m/%Y") if fecha_inicio else "-"} - {fecha_fin.strftime("%d/%m/%Y") if fecha_fin else "-"}')

        # Espacio para firmas
        y_sig = height - 260
        c.line(60, y_sig, 260, y_sig)
        c.drawString(60, y_sig - 15, 'Firma del funcionario')

        c.line(300, y_sig, 500, y_sig)
        c.drawString(300, y_sig - 15, 'Firma del jefe')

        c.line(60, y_sig - 60, 260, y_sig - 60)
        c.drawString(60, y_sig - 75, 'Firma RRHH')

        c.setFont('Helvetica-Oblique', 9)
        c.drawString(50, 40, 'Este documento es la solicitud formal del funcionario para gestionar sus vacaciones ante RRHH.')

        c.showPage()
        c.save()
        buffer.seek(0)

        filename = f'solicitud_vacaciones_{empleado.codigo}_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

    except Exception as e:
        flash(f'Error generando PDF: {e}', 'danger')
        return redirect(url_for('rrhh.listar_vacaciones'))

@rrhh_bp.route('/vacaciones/solicitar', methods=['GET', 'POST'])
@login_required
def solicitar_vacaciones():
    """Solicitar vacaciones con cálculo 100% automático y dinámico"""
    empleados = Empleado.query.filter_by(estado=EstadoEmpleadoEnum.ACTIVO).all()
    
    if request.method == 'POST':
        try:
            empleado_id = int(request.form.get('empleado_id'))
            fecha_inicio = datetime.strptime(request.form.get('fecha_inicio'), '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(request.form.get('fecha_fin'), '%Y-%m-%d').date()
            dias_solicitados = (fecha_fin - fecha_inicio).days + 1
            
            empleado = Empleado.query.get(empleado_id)
            
            # Calcular saldo actual dinámicamente
            saldo = calcular_saldo_vacaciones_actual(empleado)
            
            # Verificar si tiene días suficientes
            if saldo['pendientes'] < dias_solicitados:
                flash(f'Solo dispone de {saldo["pendientes"]} días de vacaciones disponibles. Días solicitados: {dias_solicitados}', 'danger')
                return redirect(url_for('rrhh.solicitar_vacaciones'))
            
            # Crear registro de solicitud (solo guarda cuando se solicita)
            vacacion = Vacacion(
                empleado_id=empleado_id,
                año=fecha_inicio.year,
                dias_disponibles=calcular_dias_vacaciones_por_antiguedad(empleado, fecha_inicio.year),
                dias_tomados=0,
                dias_pendientes=saldo['pendientes'],
                fecha_inicio_solicitud=fecha_inicio,
                fecha_fin_solicitud=fecha_fin,
                estado=EstadoVacacionEnum.PENDIENTE
            )
            
            db.session.add(vacacion)
            db.session.commit()
            
            registrar_operacion_crud(
                current_user, 'vacaciones', 'CREATE', 'vacaciones',
                vacacion.id, {'empleado_id': empleado_id, 'dias': dias_solicitados}
            )
            
            flash('Solicitud de vacaciones realizada exitosamente', 'success')
            return redirect(url_for('rrhh.listar_vacaciones'))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al solicitar vacaciones: {str(e)}', 'danger')
    
    return render_template('rrhh/solicitar_vacaciones.html', empleados=empleados)

@rrhh_bp.route('/vacaciones/<int:vacacion_id>/aprobar', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def aprobar_vacacion(vacacion_id):
    """Aprobar solicitud de vacaciones"""
    vacacion = Vacacion.query.get_or_404(vacacion_id)
    
    try:
        from ..models import EstadoVacacionEnum
        vacacion.estado = EstadoVacacionEnum.APROBADA
        # Actualizar días disponibles
        dias_solicitados = (vacacion.fecha_fin_solicitud - vacacion.fecha_inicio_solicitud).days + 1
        vacacion.dias_disponibles -= dias_solicitados
        vacacion.dias_tomados += dias_solicitados
        vacacion.dias_pendientes -= dias_solicitados
        db.session.commit()
        
        registrar_operacion_crud(
            current_user, 'vacaciones', 'UPDATE', 'vacaciones',
            vacacion_id, {'estado': 'APROBADA', 'dias': dias_solicitados}
        )
        
        flash(f'Vacación aprobada exitosamente ({dias_solicitados} días)', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al aprobar vacación: {str(e)}', 'danger')
    
    return redirect(url_for('rrhh.listar_vacaciones'))

@rrhh_bp.route('/vacaciones/<int:vacacion_id>/rechazar', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def rechazar_vacacion(vacacion_id):
    """Rechazar solicitud de vacaciones"""
    vacacion = Vacacion.query.get_or_404(vacacion_id)
    
    try:
        from ..models import EstadoVacacionEnum
        vacacion.estado = EstadoVacacionEnum.RECHAZADA
        db.session.commit()
        
        registrar_operacion_crud(
            current_user, 'vacaciones', 'UPDATE', 'vacaciones',
            vacacion_id, {'estado': 'RECHAZADA'}
        )
        
        flash('Solicitud de vacaciones rechazada', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al rechazar vacación: {str(e)}', 'danger')
    
    return redirect(url_for('rrhh.listar_vacaciones'))

# ==================== DESPIDOS ====================

def calcular_antiguedad_años(fecha_inicio, fecha_fin):
    """Calcula años completos entre dos fechas."""
    diferencia = fecha_fin - fecha_inicio
    años = diferencia.days // 365
    return años

def calcular_indemnizacion(salario_base, tipo_despido, antiguedad_años):
    """
    Calcula indemnización por despido.
    - Justificado: 0 (sin indemnización)
    - Injustificado: 1 mes + 1 mes por año (máx 12 meses)
    """
    if tipo_despido == 'justificado':
        return Decimal('0')
    
    # Injustificado: 1 + años (máximo 12 meses)
    meses = min(1 + antiguedad_años, 12)
    return (Decimal(meses) * Decimal(str(salario_base))).quantize(Decimal('0.01'))

def calcular_aguinaldo_proporcional(salario_base, fecha_despido, empleado_id=None):
    """
    Calcula aguinaldo proporcional (13º sueldo) según ley paraguaya.
    
    MÉTODO CORRECTO:
    - Suma TODOS los ingresos devengados en el año (salarios + extras + comisiones + bonificaciones)
    - Divide entre 12
    
    Si no hay liquidaciones registradas, usa el salario base proporcional como fallback.
    """
    año_despido = fecha_despido.year
    fecha_inicio_año = datetime(año_despido, 1, 1).date()
    
    # Si tenemos empleado_id, intentar calcular con ingresos reales
    if empleado_id:
        # Sumar todos los salarios base de las liquidaciones del año
        total_salarios = db.session.query(
            func.sum(Liquidacion.salario_base)
        ).filter(
            Liquidacion.empleado_id == empleado_id,
            Liquidacion.periodo.like(f'{año_despido}%'),
            Liquidacion.aguinaldo_monto == 0  # Excluir liquidaciones de aguinaldo
        ).scalar() or Decimal('0')
        
        # Sumar ingresos extras del año (horas extras, comisiones, bonificaciones habituales)
        total_extras = db.session.query(
            func.sum(IngresoExtra.monto)
        ).filter(
            IngresoExtra.empleado_id == empleado_id,
            IngresoExtra.año == año_despido,
            IngresoExtra.tipo.in_(['Horas Extras', 'Comisión', 'Bonificación'])  # Excluir viáticos
        ).scalar() or Decimal('0')
        
        # Total devengado = salarios + extras
        total_devengado = total_salarios + total_extras
        
        # Si hay datos reales, usar el método correcto: Total / 12
        if total_devengado > 0:
            aguinaldo = (total_devengado / Decimal('12')).quantize(Decimal('0.01'))
            return aguinaldo
    
    # FALLBACK: Si no hay liquidaciones, calcular proporcional por días trabajados
    días_trabajados = (fecha_despido - fecha_inicio_año).days + 1
    meses_trabajados = Decimal(str(días_trabajados)) / Decimal('30')
    
    aguinaldo = (Decimal(str(salario_base)) * meses_trabajados / Decimal('12')).quantize(Decimal('0.01'))
    return aguinaldo

def calcular_vacaciones_no_gozadas(empleado, fecha_despido):
    """
    Calcula monto de vacaciones no gozadas.
    - Suma vacaciones aprobadas no tomadas de años anteriores
    - Suma proporción de derechos en año actual (2 días por mes)
    """
    # Vacaciones registradas pero no completadas (asumimos que registros con estado 'aprobada' no gozadas están aquí)
    from ..models import EstadoVacacionEnum
    
    # Acumular vacaciones no gozadas del año anterior
    año_anterior = fecha_despido.year - 1
    vacaciones_año_ant = db.session.query(db.func.count(Vacacion.id)).filter(
        Vacacion.empleado_id == empleado.id,
        Vacacion.año == año_anterior,
        Vacacion.estado == EstadoVacacionEnum.APROBADA
    ).scalar() or 0
    
    # Vacaciones ganadas en año actual (2 días por mes trabajado)
    días_transcurridos = (fecha_despido - datetime(fecha_despido.year, 1, 1).date()).days + 1
    meses_año_actual = Decimal(str(días_transcurridos)) / Decimal('30')
    vacaciones_ganadas_año = meses_año_actual * Decimal('2')
    
    # Total de días no gozados
    total_días = Decimal(str(vacaciones_año_ant * 2)) + vacaciones_ganadas_año  # 2 días por mes
    
    # Valor por día = salario_base / 30
    salario_diario = Decimal(str(empleado.salario_base)) / Decimal('30')
    vacaciones_monto = (total_días * salario_diario).quantize(Decimal('0.01'))
    
    return vacaciones_monto

def calcular_aportes_ips_despido(monto_liquido, porcentaje_aporte=Decimal('0.09')):
    """
    Calcula aporte a IPS (9% del empleado).
    Se aplica al total de indemnización + aguinaldo + vacaciones.
    """
    return (Decimal(str(monto_liquido)) * porcentaje_aporte).quantize(Decimal('0.01'))

def generar_liquidacion_despido(empleado_id, tipo_despido, causal=None, descripcion=None):
    """
    Genera liquidación completa por despido.
    Retorna diccionario con montos calculados y objetos creados.
    """
    empleado = Empleado.query.get(empleado_id)
    if not empleado:
        return None
    
    fecha_despido = date.today()
    
    # Calcular antigüedad
    antiguedad = calcular_antiguedad_años(empleado.fecha_ingreso, fecha_despido)
    
    # 1. Indemnización
    indemnizacion = calcular_indemnizacion(empleado.salario_base, tipo_despido, antiguedad)
    
    # 2. Aguinaldo proporcional (método legal: suma ingresos reales del año / 12)
    aguinaldo = calcular_aguinaldo_proporcional(empleado.salario_base, fecha_despido, empleado_id=empleado.id)
    
    # 3. Vacaciones no gozadas
    vacaciones = calcular_vacaciones_no_gozadas(empleado, fecha_despido)
    
    # 4. Subtotal antes de aportes
    subtotal = indemnizacion + aguinaldo + vacaciones
    
    # 5. Aportes IPS
    aportes = calcular_aportes_ips_despido(subtotal)
    
    # 6. Total líquido
    total_liquido = subtotal - aportes
    
    # Crear registro de Despido
    despido = Despido(
        empleado_id=empleado_id,
        tipo=tipo_despido,
        causal=causal,
        descripcion=descripcion,
        fecha_despido=fecha_despido,
        usuario_id=current_user.id if current_user else None
    )
    db.session.add(despido)
    db.session.flush()
    
    # Crear registro de Liquidación
    liquidacion = Liquidacion(
        empleado_id=empleado_id,
        periodo=f"{fecha_despido.year}-{fecha_despido.month:02d}",
        salario_base=empleado.salario_base,
        salario_neto=total_liquido,
        despido_id=despido.id,
        indemnizacion_monto=indemnizacion,
        aguinaldo_monto=aguinaldo,
        vacaciones_monto=vacaciones,
        aportes_ips_despido=aportes
    )
    db.session.add(liquidacion)
    despido.liquidaciones.append(liquidacion)  # Relación inversa
    
    # Cambiar estado del empleado a INACTIVO
    empleado.estado = EstadoEmpleadoEnum.INACTIVO
    empleado.fecha_retiro = fecha_despido
    
    db.session.commit()
    
    # Registrar en bitácora
    registrar_bitacora(
        current_user,
        'despidos',
        'CREATE',
        'despidos',
        despido.id,
        f"Despido {tipo_despido}. Antigüedad: {antiguedad} años. Indemnización: {indemnizacion}"
    )
    
    return {
        'indemnizacion': indemnizacion,
        'aguinaldo': aguinaldo,
        'vacaciones': vacaciones,
        'aportes_ips': aportes,
        'total_liquido': total_liquido,
        'subtotal': subtotal,
        'antiguedad': antiguedad,
        'liquidacion_obj': liquidacion,
        'despido_obj': despido
    }

@rrhh_bp.route('/registrar_despido', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def registrar_despido():
    """Formulario y procesamiento de despido."""
    if request.method == 'POST':
        empleado_id = request.form.get('empleado_id', type=int)
        tipo = request.form.get('tipo')
        causal = request.form.get('causal', default=None)
        descripcion = request.form.get('descripcion', default=None)
        
        resultado = generar_liquidacion_despido(empleado_id, tipo, causal, descripcion)
        
        if resultado:
            flash('Despido registrado. Liquidación generada.', 'success')
            return redirect(url_for('rrhh.ver_liquidacion_despido', liquidacion_id=resultado['liquidacion_obj'].id))
        else:
            flash('Error al registrar despido.', 'danger')
    
    empleados = Empleado.query.filter_by(estado=EstadoEmpleadoEnum.ACTIVO).all()
    causales_justificadas = [
        ('incapacidad', 'Incapacidad Laboral Permanente'),
        ('ineptitud', 'Ineptitud Técnica o Manifiesta'),
        ('falta_grave', 'Falta Grave / Conducta Inapropiada'),
        ('perdida_habilitacion', 'Pérdida de Habilitación Profesional'),
        ('fuerza_mayor', 'Fuerza Mayor o Caso Fortuito'),
    ]
    
    return render_template(
        'rrhh/registrar_despido.html',
        empleados=empleados,
        causales=causales_justificadas
    )

@rrhh_bp.route('/liquidacion_despido/<int:liquidacion_id>')
@login_required
@role_required(RoleEnum.RRHH)
def ver_liquidacion_despido(liquidacion_id):
    """Vista de detalles de liquidación por despido."""
    liquidacion = Liquidacion.query.get_or_404(liquidacion_id)
    if not liquidacion.despido_id:
        flash('Esta liquidación no es de despido.', 'warning')
        return redirect(url_for('rrhh.liquidaciones'))
    
    despido = liquidacion.despido
    empleado = liquidacion.empleado
    antiguedad = calcular_antiguedad_años(empleado.fecha_ingreso, despido.fecha_despido)
    
    return render_template(
        'rrhh/liquidacion_despido.html',
        liquidacion=liquidacion,
        despido=despido,
        empleado=empleado,
        antiguedad=antiguedad
    )

@rrhh_bp.route('/liquidacion_despido/<int:liquidacion_id>/descargar_pdf')
@login_required
@role_required(RoleEnum.RRHH)
def descargar_pdf_liquidacion_despido(liquidacion_id):
    """Descarga PDF de liquidación por despido."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    
    liquidacion = Liquidacion.query.get_or_404(liquidacion_id)
    despido = liquidacion.despido
    empleado = liquidacion.empleado
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch,
        leftMargin=0.75*inch,
        rightMargin=0.75*inch
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = styles['Title']
    title_style.alignment = 1  # Center
    elements.append(Paragraph("LIQUIDACIÓN POR DESPIDO", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Subtítulo
    elements.append(Paragraph(f"<b>Fecha:</b> {despido.fecha_despido.strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Spacer(1, 0.15*inch))
    
    # Datos empleado
    data = [
        ['EMPLEADO:', empleado.nombre],
        ['CÉDULA:', empleado.ci],
        ['CARGO:', empleado.cargo.nombre if empleado.cargo else '---'],
        ['ANTIGÜEDAD:', f"{calcular_antiguedad_años(empleado.fecha_ingreso, despido.fecha_despido)} años"],
        ['TIPO DESPIDO:', despido.tipo.upper()],
        ['CAUSAL:', despido.causal.upper() if despido.causal else '---'],
    ]
    
    table = Table(data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 9),
        ('FONT', (1, 0), (1, -1), 'Helvetica', 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), HexColor('#e8f4f8')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.25*inch))
    
    # Tabla de rubros
    rubros = [
        ['RUBRO', 'MONTO'],
        ['Indemnización por antigüedad', f"${float(liquidacion.indemnizacion_monto):,.2f}"],
        ['Aguinaldo (13º Sueldo) proporcional', f"${float(liquidacion.aguinaldo_monto):,.2f}"],
        ['Vacaciones no gozadas', f"${float(liquidacion.vacaciones_monto):,.2f}"],
        ['Subtotal', f"${float(liquidacion.indemnizacion_monto + liquidacion.aguinaldo_monto + liquidacion.vacaciones_monto):,.2f}"],
        ['(-) Aporte IPS (9%)', f"-${float(liquidacion.aportes_ips_despido):,.2f}"],
        ['TOTAL NETO A PAGAR', f"${float(liquidacion.salario_neto):,.2f}"],
    ]
    
    tabla_rubros = Table(rubros, colWidths=[3.5*inch, 2*inch])
    tabla_rubros.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
        ('FONT', (0, 1), (-1, -2), 'Helvetica', 9),
        ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 11),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, -1), (-1, -1), HexColor('#D9E1F2')),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(tabla_rubros)
    elements.append(Spacer(1, 0.3*inch))
    
    # Observaciones
    if despido.descripcion:
        elements.append(Paragraph("<b>Observaciones:</b>", styles['Normal']))
        elements.append(Paragraph(despido.descripcion, styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f"Liquidacion_Despido_{empleado.nombre.replace(' ', '_')}_{despido.fecha_despido.strftime('%Y%m%d')}.pdf"
    )

# ==================== BITACORA ====================
@rrhh_bp.route('/bitacora', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def ver_bitacora():
    """Ver bitácora de acciones"""
    from ..models import Usuario
    
    pagina = request.args.get('page', 1, type=int)
    usuario_filtro = request.args.get('usuario', '')
    modulo_filtro = request.args.get('modulo', '')
    
    query = Bitacora.query
    
    if usuario_filtro:
        query = query.filter(Bitacora.usuario.has(Usuario.nombre_usuario.ilike(f'%{usuario_filtro}%')))
    
    if modulo_filtro:
        query = query.filter(Bitacora.modulo == modulo_filtro)
    
    bitacoras = query.order_by(desc(Bitacora.fecha_creacion)).paginate(page=pagina, per_page=20)
    
    return render_template('rrhh/bitacora.html', bitacoras=bitacoras)

# ==================== AGUINALDOS ====================

def generar_aguinaldos_anual(año, mes_corte=None, día_corte=None):
    """
    Genera aguinaldos anuales para todos los empleados activos.
    
    Args:
        año: Año para el cual se calcula el aguinaldo (ej: 2025)
        mes_corte: Mes de pago (por defecto 12 - diciembre)
        día_corte: Día de pago (por defecto 31)
    
    Returns:
        Diccionario con resumen de generación
    """
    if mes_corte is None:
        mes_corte = 12
    if día_corte is None:
        día_corte = 31
    
    # Fecha de corte para cálculo
    fecha_corte = date(año, mes_corte, día_corte)
    
    # Obtener empleados activos
    empleados = Empleado.query.filter_by(estado=EstadoEmpleadoEnum.ACTIVO).all()
    
    generados = 0
    duplicados = 0
    errores = 0
    total_bruto = Decimal('0')
    total_neto = Decimal('0')
    
    for empleado in empleados:
        try:
            # Verificar si ya existe aguinaldo para este año
            liquidacion_existente = Liquidacion.query.filter(
                Liquidacion.empleado_id == empleado.id,
                Liquidacion.periodo.like(f'{año}%'),
                Liquidacion.aguinaldo_monto > 0  # Solo si ya hay aguinaldo
            ).first()
            
            if liquidacion_existente:
                duplicados += 1
                continue
            
            # Calcular aguinaldo proporcional para el año
            fecha_inicio_año = date(año, 1, 1)
            
            # Si el empleado fue contratado después del inicio del año, usar fecha de contratación
            fecha_desde = max(fecha_inicio_año, empleado.fecha_ingreso)
            
            # Si el empleado se retiró antes del corte, usar fecha de retiro
            fecha_hasta = fecha_corte
            if empleado.fecha_retiro and empleado.fecha_retiro < fecha_corte:
                fecha_hasta = empleado.fecha_retiro
            
            # Calcular días trabajados en el año
            días_trabajados = (fecha_hasta - fecha_desde).days + 1
            if días_trabajados <= 0:
                continue
            
            # CÁLCULO SEGÚN LEY PARAGUAYA:
            # Sumar todos los salarios base de liquidaciones del año
            total_salarios = db.session.query(
                func.sum(Liquidacion.salario_base)
            ).filter(
                Liquidacion.empleado_id == empleado.id,
                Liquidacion.periodo.like(f'{año}%'),
                Liquidacion.aguinaldo_monto == 0  # Excluir liquidaciones de aguinaldo previas
            ).scalar() or Decimal('0')
            
            # Sumar ingresos extras del año (horas extras, comisiones, bonificaciones)
            total_extras = db.session.query(
                func.sum(IngresoExtra.monto)
            ).filter(
                IngresoExtra.empleado_id == empleado.id,
                IngresoExtra.año == año,
                IngresoExtra.tipo.in_(['Horas Extras', 'Comisión', 'Bonificación'])  # Excluir viáticos
            ).scalar() or Decimal('0')
            
            # Total devengado en el año
            total_devengado = total_salarios + total_extras
            
            # Si hay liquidaciones registradas, usar el método correcto: Total / 12
            if total_devengado > 0:
                aguinaldo_bruto = (total_devengado / Decimal('12')).quantize(Decimal('0.01'))
            else:
                # FALLBACK: Si no hay liquidaciones, calcular proporcional por meses trabajados
                meses_trabajados = Decimal(str(días_trabajados)) / Decimal('30')
                aguinaldo_bruto = (Decimal(str(empleado.salario_base)) * meses_trabajados / Decimal('12')).quantize(Decimal('0.01'))
            
            # Aportes IPS = 9% sobre aguinaldo
            aportes_ips = (aguinaldo_bruto * Decimal('0.09')).quantize(Decimal('0.01'))
            
            # Aguinaldo neto
            aguinaldo_neto = (aguinaldo_bruto - aportes_ips).quantize(Decimal('0.01'))
            
            # Crear liquidación de aguinaldo
            periodo_aguinaldo = f"{año}-{mes_corte:02d}"
            
            liquidacion = Liquidacion(
                empleado_id=empleado.id,
                periodo=periodo_aguinaldo,
                salario_base=empleado.salario_base,
                salario_neto=aguinaldo_neto,
                aguinaldo_monto=aguinaldo_bruto,
                aportes_ips_despido=aportes_ips,
                indemnizacion_monto=Decimal('0'),
                vacaciones_monto=Decimal('0'),
                ingresos_extras=Decimal('0'),
                descuentos=Decimal('0'),
                aporte_ips=Decimal('0'),
                dias_trabajados=días_trabajados
            )
            
            db.session.add(liquidacion)
            db.session.flush()
            
            generados += 1
            total_bruto += aguinaldo_bruto
            total_neto += aguinaldo_neto
            
        except Exception as e:
            errores += 1
            db.session.rollback()
            continue
    
    try:
        db.session.commit()
        registrar_bitacora(
            current_user,
            'aguinaldos',
            'CREATE',
            'liquidaciones',
            detalle=f'Aguinaldos {año} generados: {generados} empleados. Total bruto: {total_bruto}. Total neto: {total_neto}'
        )
    except Exception as e:
        db.session.rollback()
        return {
            'success': False,
            'error': str(e),
            'generados': 0
        }
    
    return {
        'success': True,
        'generados': generados,
        'duplicados': duplicados,
        'errores': errores,
        'total_bruto': total_bruto,
        'total_neto': total_neto,
        'fecha_corte': fecha_corte.strftime('%d/%m/%Y')
    }

@rrhh_bp.route('/aguinaldos', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def listar_aguinaldos():
    """Lista aguinaldos generados por año"""
    registrar_bitacora(current_user, 'aguinaldos', 'VIEW', 'liquidaciones')
    
    año_filtro = request.args.get('año', date.today().year, type=int)
    pagina = request.args.get('page', 1, type=int)
    
    # Obtener liquidaciones que contengan aguinaldo (aguinaldo_monto > 0)
    aguinaldos = Liquidacion.query.filter(
        Liquidacion.aguinaldo_monto > 0,
        Liquidacion.periodo.like(f'{año_filtro}%')
    ).order_by(desc(Liquidacion.fecha_generacion)).paginate(page=pagina, per_page=15)
    
    # Calcular totales para el año
    totales = db.session.query(
        func.sum(Liquidacion.aguinaldo_monto).label('total_bruto'),
        func.sum(Liquidacion.aportes_ips_despido).label('total_ips'),
        func.sum(Liquidacion.salario_neto).label('total_neto'),
        func.count(Liquidacion.id).label('cantidad')
    ).filter(
        Liquidacion.aguinaldo_monto > 0,
        Liquidacion.periodo.like(f'{año_filtro}%')
    ).first()
    
    return render_template(
        'rrhh/aguinaldos_listado.html',
        aguinaldos=aguinaldos,
        año=año_filtro,
        totales=totales or {
            'total_bruto': Decimal('0'),
            'total_ips': Decimal('0'),
            'total_neto': Decimal('0'),
            'cantidad': 0
        }
    )

@rrhh_bp.route('/generar_aguinaldos', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def generar_aguinaldos():
    """Formulario y procesamiento para generar aguinaldos anuales"""
    resultado = None
    
    if request.method == 'POST':
        año = int(request.form.get('año', date.today().year))
        mes_corte = int(request.form.get('mes_corte', 12))
        día_corte = int(request.form.get('día_corte', 31))
        acción = request.form.get('acción', 'preview')
        
        # Contar empleados activos que recibirían aguinaldo
        empleados_activos = Empleado.query.filter_by(estado=EstadoEmpleadoEnum.ACTIVO).count()
        
        if acción == 'preview':
            # Solo mostrar preview (preview en el mismo formulario)
            fecha_corte = date(año, mes_corte, día_corte)
            preview_datos = []
            
            empleados = Empleado.query.filter_by(estado=EstadoEmpleadoEnum.ACTIVO).all()
            
            for empleado in empleados:
                fecha_desde = max(date(año, 1, 1), empleado.fecha_ingreso)
                fecha_hasta = fecha_corte
                if empleado.fecha_retiro and empleado.fecha_retiro < fecha_corte:
                    fecha_hasta = empleado.fecha_retiro
                
                días = (fecha_hasta - fecha_desde).days + 1
                meses = Decimal(str(días)) / Decimal('30')
                
                # CÁLCULO SEGÚN LEY PARAGUAYA:
                # Sumar todos los salarios base de liquidaciones del año
                total_salarios = db.session.query(
                    func.sum(Liquidacion.salario_base)
                ).filter(
                    Liquidacion.empleado_id == empleado.id,
                    Liquidacion.periodo.like(f'{año}%'),
                    Liquidacion.aguinaldo_monto == 0
                ).scalar() or Decimal('0')
                
                # Sumar ingresos extras del año
                total_extras = db.session.query(
                    func.sum(IngresoExtra.monto)
                ).filter(
                    IngresoExtra.empleado_id == empleado.id,
                    IngresoExtra.año == año,
                    IngresoExtra.tipo.in_(['Horas Extras', 'Comisión', 'Bonificación'])
                ).scalar() or Decimal('0')
                
                # Total devengado
                total_devengado = total_salarios + total_extras
                
                # Si hay datos reales, usar método correcto: Total / 12
                if total_devengado > 0:
                    aguinaldo = (total_devengado / Decimal('12')).quantize(Decimal('0.01'))
                else:
                    # FALLBACK: usar proporcional por meses trabajados
                    aguinaldo = (Decimal(str(empleado.salario_base)) * meses / Decimal('12')).quantize(Decimal('0.01'))
                
                ips = (aguinaldo * Decimal('0.09')).quantize(Decimal('0.01'))
                neto = (aguinaldo - ips).quantize(Decimal('0.01'))
                
                preview_datos.append({
                    'empleado': empleado,
                    'meses': meses,
                    'total_devengado': total_devengado,
                    'aguinaldo_bruto': aguinaldo,
                    'ips': ips,
                    'aguinaldo_neto': neto
                })
            
            resultado = {
                'tipo': 'preview',
                'año': año,
                'fecha_corte': fecha_corte.strftime('%d/%m/%Y'),
                'datos': preview_datos,
                'total_bruto': sum(d['aguinaldo_bruto'] for d in preview_datos),
                'total_ips': sum(d['ips'] for d in preview_datos),
                'total_neto': sum(d['aguinaldo_neto'] for d in preview_datos),
                'cantidad': len(preview_datos)
            }
        
        elif acción == 'generar':
            # Generar aguinaldos reales
            resultado = generar_aguinaldos_anual(año, mes_corte, día_corte)
            resultado['tipo'] = 'generado'
            
            if resultado['success']:
                flash(
                    f"✓ {resultado['generados']} aguinaldos generados. Total: {resultado['total_neto']:,.2f} Gs.",
                    'success'
                )
                return redirect(url_for('rrhh.listar_aguinaldos', año=año))
            else:
                flash(f"Error al generar aguinaldos: {resultado.get('error', 'Error desconocido')}", 'danger')
    
    años_disponibles = list(range(date.today().year, 2020, -1))
    
    return render_template(
        'rrhh/generar_aguinaldos.html',
        resultado=resultado,
        años_disponibles=años_disponibles,
        año_actual=date.today().year
    )

# ==================== LEGAJO DIGITAL / PERFIL EMPLEADO ====================

@rrhh_bp.route('/empleados/<int:empleado_id>/perfil', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def perfil_empleado(empleado_id):
    """Muestra el perfil/legajo completo del empleado"""
    empleado = Empleado.query.get_or_404(empleado_id)
    registrar_bitacora(current_user, 'empleados', 'VIEW', 'empleados', empleado_id, 'ver_perfil')
    # Si el empleado está Inactivo, habilitar modo solo-lectura en la vista
    read_only = (empleado.estado == EstadoEmpleadoEnum.INACTIVO)
    
    # 📊 Calcular estadísticas de justificaciones del año actual
    año_actual = date.today().year
    ausencias_justificadas = db.session.query(func.count(Asistencia.id)).filter(
        Asistencia.empleado_id == empleado_id,
        func.extract('year', Asistencia.fecha) == año_actual,
        Asistencia.presente == False,
        Asistencia.justificacion_estado == 'JUSTIFICADO'
    ).scalar() or 0
    
    ausencias_injustificadas = db.session.query(func.count(Asistencia.id)).filter(
        Asistencia.empleado_id == empleado_id,
        func.extract('year', Asistencia.fecha) == año_actual,
        Asistencia.presente == False,
        Asistencia.justificacion_estado == 'INJUSTIFICADO'
    ).scalar() or 0
    
    ausencias_pendientes = db.session.query(func.count(Asistencia.id)).filter(
        Asistencia.empleado_id == empleado_id,
        func.extract('year', Asistencia.fecha) == año_actual,
        Asistencia.presente == False,
        Asistencia.justificacion_estado == 'PENDIENTE'
    ).scalar() or 0

    return render_template('rrhh/empleado_perfil.html', 
                         empleado=empleado, 
                         read_only=read_only,
                         ausencias_justificadas=ausencias_justificadas,
                         ausencias_injustificadas=ausencias_injustificadas,
                         ausencias_pendientes=ausencias_pendientes,
                         date=date)

@rrhh_bp.route('/api/empleados/<int:empleado_id>/general', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def api_empleado_general(empleado_id):
    """API: Datos generales y KPIs del empleado"""
    empleado = Empleado.query.get_or_404(empleado_id)
    
    # Calcular datos
    vacaciones_total = db.session.query(func.sum(Vacacion.dias_disponibles)).filter_by(empleado_id=empleado_id).scalar() or 0
    vacaciones_tomadas = db.session.query(func.sum(Vacacion.dias_tomados)).filter_by(empleado_id=empleado_id).scalar() or 0
    vacaciones_pendientes = vacaciones_total - vacaciones_tomadas
    
    sanciones_activas = Sancion.query.filter_by(empleado_id=empleado_id).count()
    
    # Permisos usados: contar los que fueron aprobados o completados
    permisos_usados = Permiso.query.filter(
        Permiso.empleado_id == empleado_id,
        or_(Permiso.estado == EstadoPermisoEnum.APROBADO, Permiso.estado == EstadoPermisoEnum.COMPLETADO)
    ).count()
    
    # Asistencia del mes actual
    mes_actual = date.today().month
    año_actual = date.today().year
    asistencias_mes = db.session.query(func.count(Asistencia.id)).filter(
        Asistencia.empleado_id == empleado_id,
        func.extract('month', Asistencia.fecha) == mes_actual,
        func.extract('year', Asistencia.fecha) == año_actual,
        Asistencia.presente == True
    ).scalar() or 0
    
    dias_habiles_mes = len([d for d in calendar.monthcalendar(año_actual, mes_actual) 
                           for dow in d if dow != 0 and datetime(año_actual, mes_actual, dow).weekday() < 5])
    
    porcentaje_asistencia = (asistencias_mes / dias_habiles_mes * 100) if dias_habiles_mes > 0 else 0
    
    # Resumen de asistencia de hoy
    resumen_hoy = _resumir_dia_asistencias(empleado_id, date.today())
    # Ingresos extras recientes (últimos 5)
    try:
        ingresos_q = IngresoExtra.query.filter_by(empleado_id=empleado_id).order_by(IngresoExtra.fecha_creacion.desc()).limit(5).all()
    except Exception:
        ingresos_q = []
    ingresos_list = []
    for ie in ingresos_q:
        ingresos_list.append({
            'id': ie.id,
            'tipo': getattr(ie, 'tipo', None),
            'monto': float(ie.monto) if getattr(ie, 'monto', None) is not None else None,
            'mes': getattr(ie, 'mes', None),
            'año': getattr(ie, 'año', None),
            'estado': getattr(ie, 'estado', None),
            'justificativo': getattr(ie, 'justificativo_archivo', None)
        })
    return jsonify({
        'id': empleado.id,
        'codigo': empleado.codigo,
        'nombre_completo': empleado.nombre_completo,
        'cargo': empleado.cargo.nombre,
        'email': empleado.email,
        'telefono': empleado.telefono,
        'fecha_ingreso': empleado.fecha_ingreso.strftime('%d/%m/%Y') if empleado.fecha_ingreso else 'N/A',
        'antiguedad': empleado.antiguedad_texto,
        'estado': empleado.estado.value,
        'salario_base': float(empleado.salario_base),
        'vacaciones_pendientes': int(vacaciones_pendientes),
        'vacaciones_total': int(vacaciones_total),
        'sanciones_activas': sanciones_activas,
        'permisos_usados': permisos_usados,
        'asistencia_mes': f"{porcentaje_asistencia:.1f}%",
        'asistencias_count': asistencias_mes,
        'resumen_hoy': resumen_hoy
    })

@rrhh_bp.route('/api/empleados/<int:empleado_id>/asistencias', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def api_empleado_asistencias(empleado_id):
    """API: Asistencias del empleado con paginación y filtros"""
    empleado = Empleado.query.get_or_404(empleado_id)
    
    mes = request.args.get('mes', type=int)
    año = request.args.get('year', type=int)
    pagina = request.args.get('page', 1, type=int)
    por_pagina = request.args.get('per_page', 20, type=int)
    
    query = Asistencia.query.filter_by(empleado_id=empleado_id).order_by(desc(Asistencia.fecha))
    
    if mes and año:
        query = query.filter(
            func.extract('month', Asistencia.fecha) == mes,
            func.extract('year', Asistencia.fecha) == año
        )
    
    paginated = query.paginate(page=pagina, per_page=por_pagina)
    
    asistencias = [{
        'id': a.id,
        'fecha': a.fecha.strftime('%d/%m/%Y'),
        'entrada': a.hora_entrada.strftime('%H:%M') if a.hora_entrada else '-',
        'salida': a.hora_salida.strftime('%H:%M') if a.hora_salida else '-',
        'presente': 'Sí' if a.presente else 'No',
        'observaciones': a.observaciones or '-'
    } for a in paginated.items]
    
    return jsonify({
        'items': asistencias,
        'total': paginated.total,
        'pages': paginated.pages,
        'current_page': pagina,
        'per_page': por_pagina
    })

@rrhh_bp.route('/api/empleados/<int:empleado_id>/vacaciones', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def api_empleado_vacaciones(empleado_id):
    """API: Histórico de vacaciones"""
    empleado = Empleado.query.get_or_404(empleado_id)
    
    vacaciones = Vacacion.query.filter_by(empleado_id=empleado_id).order_by(desc(Vacacion.año)).all()
    
    datos = [{
        'año': v.año,
        'dias_disponibles': v.dias_disponibles,
        'dias_tomados': v.dias_tomados,
        'dias_pendientes': v.dias_pendientes,
        'fecha_inicio': v.fecha_inicio_solicitud.strftime('%d/%m/%Y') if v.fecha_inicio_solicitud else '-',
        'fecha_fin': v.fecha_fin_solicitud.strftime('%d/%m/%Y') if v.fecha_fin_solicitud else '-',
        'estado': v.estado.value if v.estado else 'N/A'
    } for v in vacaciones]
    
    return jsonify({
        'items': datos,
        'total': len(datos)
    })

@rrhh_bp.route('/api/empleados/<int:empleado_id>/justificaciones', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def api_empleado_justificaciones(empleado_id):
    """API: Historial de ausencias con justificaciones"""
    empleado = Empleado.query.get_or_404(empleado_id)
    
    mes = request.args.get('mes', type=int)
    año = request.args.get('year', date.today().year, type=int)
    pagina = request.args.get('page', 1, type=int)
    por_pagina = request.args.get('per_page', 20, type=int)
    
    query = Asistencia.query.filter(
        Asistencia.empleado_id == empleado_id,
        Asistencia.presente == False
    ).order_by(desc(Asistencia.fecha))
    
    if mes:
        query = query.filter(func.extract('month', Asistencia.fecha) == mes)
    if año:
        query = query.filter(func.extract('year', Asistencia.fecha) == año)
    
    paginated = query.paginate(page=pagina, per_page=por_pagina)
    
    ausencias = []
    for a in paginated.items:
        justificador = None
        if a.justificacion_por:
            user = Usuario.query.get(a.justificacion_por)
            justificador = user.username if user else None
        
        ausencias.append({
            'id': a.id,
            'fecha': a.fecha.strftime('%d/%m/%Y'),
            'estado': a.justificacion_estado or 'PENDIENTE',
            'nota': a.justificacion_nota or '',
            'justificado_por': justificador,
            'fecha_justificacion': a.justificacion_fecha.strftime('%d/%m/%Y %H:%M') if a.justificacion_fecha else None,
            'observaciones': a.observaciones or ''
        })
    
    return jsonify({
        'items': ausencias,
        'total': paginated.total,
        'pages': paginated.pages,
        'current_page': pagina,
        'per_page': por_pagina
    })

@rrhh_bp.route('/api/empleados/<int:empleado_id>/permisos', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def api_empleado_permisos(empleado_id):
    """API: Permisos del empleado con paginación"""
    empleado = Empleado.query.get_or_404(empleado_id)
    
    pagina = request.args.get('page', 1, type=int)
    por_pagina = request.args.get('per_page', 15, type=int)
    
    query = Permiso.query.filter_by(empleado_id=empleado_id).order_by(desc(Permiso.fecha_inicio))
    paginated = query.paginate(page=pagina, per_page=por_pagina)
    
    permisos = [{
        'id': p.id,
        'tipo': p.tipo_permiso,
        'motivo': p.motivo,
        'fecha_inicio': p.fecha_inicio.strftime('%d/%m/%Y'),
        'fecha_fin': p.fecha_fin.strftime('%d/%m/%Y'),
        'dias': p.dias_solicitados or ((p.fecha_fin - p.fecha_inicio).days + 1),
        'estado': p.estado.value if p.estado else 'N/A',
        'con_goce': 'Sí' if p.con_goce else 'No',
        'justificativo': p.justificativo_archivo if getattr(p, 'justificativo_archivo', None) else None
    } for p in paginated.items]
    
    return jsonify({
        'items': permisos,
        'total': paginated.total,
        'pages': paginated.pages,
        'current_page': pagina,
        'per_page': por_pagina
    })

@rrhh_bp.route('/api/empleados/<int:empleado_id>/sanciones', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def api_empleado_sanciones(empleado_id):
    """API: Sanciones del empleado con paginación"""
    empleado = Empleado.query.get_or_404(empleado_id)
    
    pagina = request.args.get('page', 1, type=int)
    por_pagina = request.args.get('per_page', 15, type=int)
    
    query = Sancion.query.filter_by(empleado_id=empleado_id).order_by(desc(Sancion.fecha))
    paginated = query.paginate(page=pagina, per_page=por_pagina)
    
    sanciones = [{
        'id': s.id,
        'tipo': s.tipo_sancion,
        'motivo': s.motivo,
        'monto': float(s.monto),
        'fecha': s.fecha.strftime('%d/%m/%Y'),
        'descripcion': s.descripcion or '-',
        'justificativo': s.justificativo_archivo if getattr(s, 'justificativo_archivo', None) else None
    } for s in paginated.items]
    
    return jsonify({
        'items': sanciones,
        'total': paginated.total,
        'pages': paginated.pages,
        'current_page': pagina,
        'per_page': por_pagina
    })

@rrhh_bp.route('/api/empleados/<int:empleado_id>/liquidaciones', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def api_empleado_liquidaciones(empleado_id):
    """API: Histórico de liquidaciones con paginación"""
    empleado = Empleado.query.get_or_404(empleado_id)
    
    pagina = request.args.get('page', 1, type=int)
    por_pagina = request.args.get('per_page', 12, type=int)
    
    query = Liquidacion.query.filter_by(empleado_id=empleado_id).order_by(desc(Liquidacion.periodo))
    paginated = query.paginate(page=pagina, per_page=por_pagina)
    
    liquidaciones = [{
        'id': l.id,
        'periodo': l.periodo,
        'salario_base': float(l.salario_base),
        'ingresos_extras': float(l.ingresos_extras),
        'descuentos': float(l.descuentos),
        'aporte_ips': float(l.aporte_ips),
        'salario_neto': float(l.salario_neto),
        'dias_trabajados': l.dias_trabajados or 30
    } for l in paginated.items]
    
    return jsonify({
        'items': liquidaciones,
        'total': paginated.total,
        'pages': paginated.pages,
        'current_page': pagina,
        'per_page': por_pagina
    })


# ===================== UPLOADS: Permisos y Sanciones =====================
@rrhh_bp.route('/permisos/<int:permiso_id>/upload-justificativo', methods=['POST'])
@login_required
def upload_permiso_justificativo(permiso_id):
    """Recibe un multipart/form-data con campo 'file' y lo asocia al permiso"""
    print(f"\n=== DEBUG upload_permiso_justificativo ===")
    print(f"Permiso ID: {permiso_id}")
    print(f"request.files keys: {list(request.files.keys())}")
    print(f"request.method: {request.method}")
    print(f"request.content_type: {request.content_type}")
    
    permiso = Permiso.query.get_or_404(permiso_id)

    # aceptar 'file' o 'justificativo' como nombre de campo para compatibilidad
    if 'file' in request.files:
        file = request.files['file']
        print(f"Archivo encontrado en 'file'")
    elif 'justificativo' in request.files:
        file = request.files['justificativo']
        print(f"Archivo encontrado en 'justificativo'")
    else:
        print(f"ERROR: No se encontró archivo en request.files")
        return jsonify({'error': 'No se recibió ningún archivo (campo "file" o "justificativo" esperado)'}), 400

    if not file or file.filename == '':
        print(f"ERROR: file vacío o sin nombre")
        return jsonify({'error': 'No se seleccionó ningún archivo'}), 400

    filename_clean = secure_filename(file.filename)
    if not allowed_file(filename_clean):
        print(f"ERROR: Extensión no permitida: {filename_clean}")
        return jsonify({'error': 'Extensión de archivo no permitida'}), 400

    os.makedirs(PERMISOS_UPLOAD_FOLDER, exist_ok=True)
    filename = secure_filename(f"perm_{permiso_id}_{int(datetime.utcnow().timestamp())}_{filename_clean}")
    filepath = os.path.join(PERMISOS_UPLOAD_FOLDER, filename)
    file.save(filepath)

    # Guardar solo "permisos/archivo.pdf" porque serve_uploads ya agrega "uploads/"
    permiso.justificativo_archivo = f"permisos/{filename}"
    db.session.commit()
    
    print(f"✓ Archivo guardado - Permiso {permiso_id}, ruta: {permiso.justificativo_archivo}, archivo existe: {os.path.exists(filepath)}")

    registrar_operacion_crud(current_user, 'permisos', 'UPDATE', 'permisos', permiso.id, {'justificativo': permiso.justificativo_archivo})

    return jsonify({'message': 'uploaded', 'ruta': permiso.justificativo_archivo}), 200


@rrhh_bp.route('/sanciones/<int:sancion_id>/upload-justificativo', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def upload_sancion_justificativo(sancion_id):
    """Recibe un multipart/form-data con campo 'file' y lo asocia a la sancion"""
    sancion = Sancion.query.get_or_404(sancion_id)

    # aceptar 'file' o 'justificativo' como nombre de campo para compatibilidad
    if 'file' in request.files:
        file = request.files['file']
    elif 'justificativo' in request.files:
        file = request.files['justificativo']
    else:
        return jsonify({'error': 'No se recibió ningún archivo (campo "file" o "justificativo" esperado)'}), 400

    if not file or file.filename == '':
        return jsonify({'error': 'No se seleccionó ningún archivo'}), 400

    filename_clean = secure_filename(file.filename)
    if not allowed_file(filename_clean):
        return jsonify({'error': 'Extensión de archivo no permitida'}), 400

    os.makedirs(SANCIONES_UPLOAD_FOLDER, exist_ok=True)
    filename = secure_filename(f"sanc_{sancion_id}_{int(datetime.utcnow().timestamp())}_{filename_clean}")
    filepath = os.path.join(SANCIONES_UPLOAD_FOLDER, filename)
    file.save(filepath)

    sancion.justificativo_archivo = os.path.relpath(filepath, os.path.dirname(os.path.dirname(__file__)))
    db.session.commit()

    registrar_operacion_crud(current_user, 'sanciones', 'UPDATE', 'sanciones', sancion.id, {'justificativo': sancion.justificativo_archivo})

    return jsonify({'message': 'uploaded', 'ruta': sancion.justificativo_archivo}), 200


# ===================== MÓDULO POSTULANTES (RECLUTAMIENTO) =====================

@rrhh_bp.route('/anticipos/create', methods=['POST'])
@login_required
def crear_anticipo():
    """Crear un anticipo para un empleado"""
    print("=== INICIO crear_anticipo ===")
    print(f"Content-Type: {request.content_type}")
    print(f"Method: {request.method}")
    print(f"Data: {request.data}")
    print(f"Is JSON: {request.is_json}")
    
    try:
        # Obtener datos de JSON
        data = request.get_json(force=True, silent=False)
        print(f"JSON data parseado: {data}")
        
        empleado_id = data.get('empleado_id')
        monto = data.get('monto')
        observaciones = data.get('observaciones')

        print(f"Parámetros: empleado_id={empleado_id}, monto={monto}, observaciones={observaciones}")

        if not empleado_id or monto is None:
            print("ERROR: Faltan parámetros requeridos")
            return jsonify({'error': 'empleado_id y monto son requeridos'}), 400

        empleado = Empleado.query.get(int(empleado_id))
        if not empleado:
            print(f"ERROR: Empleado {empleado_id} no encontrado")
            return jsonify({'error': 'Empleado no encontrado'}), 404

        # No permitir anticipos mayores al salario base
        monto_dec = Decimal(str(monto))
        print(f"Monto: {monto_dec}, Salario base: {empleado.salario_base}")
        
        if monto_dec > empleado.salario_base:
            print(f"ERROR: Anticipo mayor que salario")
            return jsonify({'error': 'El anticipo no puede ser mayor al salario base'}), 400

        anticipo = Anticipo(
            empleado_id=empleado.id,
            monto=monto_dec,
            observaciones=observaciones
        )
        db.session.add(anticipo)
        db.session.commit()
        print(f"✓ Anticipo creado: ID={anticipo.id}")

        registrar_operacion_crud(current_user, 'anticipos', 'CREATE', 'anticipos', anticipo.id, {'empleado_id': empleado.id, 'monto': str(monto_dec)})

        return jsonify({'message': 'anticipo creado', 'id': anticipo.id}), 200
    except Exception as e:
        db.session.rollback()
        print(f"❌ EXCEPTION: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@rrhh_bp.route('/api/empleados/<int:empleado_id>/anticipos', methods=['GET'])
@login_required
def api_empleado_anticipos(empleado_id):
    """API: listar anticipos de un empleado"""
    empleado = Empleado.query.get_or_404(empleado_id)
    pagina = request.args.get('page', 1, type=int)
    por_pagina = request.args.get('per_page', 12, type=int)
    query = Anticipo.query.filter_by(empleado_id=empleado_id).order_by(desc(Anticipo.fecha_solicitud))
    paginated = query.paginate(page=pagina, per_page=por_pagina)

    items = [{
        'id': a.id,
        'monto': float(a.monto),
        'fecha_solicitud': a.fecha_solicitud.strftime('%d/%m/%Y %H:%M:%S'),
        'aprobado': a.aprobado,
        'rechazado': getattr(a, 'rechazado', False),
        'fecha_rechazo': a.fecha_rechazo.strftime('%d/%m/%Y %H:%M:%S') if getattr(a, 'fecha_rechazo', None) else None,
        'aplicado': a.aplicado,
        'justificativo': a.justificativo_archivo
    } for a in paginated.items]

    return jsonify({'items': items, 'total': paginated.total, 'pages': paginated.pages, 'current_page': pagina, 'per_page': por_pagina})


@rrhh_bp.route('/anticipos/<int:anticipo_id>/approve', methods=['POST'])
@login_required
def aprobar_anticipo(anticipo_id):
    """Aprobar anticipo y generar PDF de recibo"""
    anticipo = Anticipo.query.get_or_404(anticipo_id)
    try:
        anticipo.aprobado = True
        anticipo.aprobado_por = current_user.id
        anticipo.fecha_aprobacion = datetime.utcnow()
        db.session.commit()

        # Generar PDF simple con reportlab
        os.makedirs(ANTICIPOS_UPLOAD_FOLDER, exist_ok=True)
        filename = secure_filename(f"anticipo_{anticipo_id}_{int(datetime.utcnow().timestamp())}.pdf")
        filepath = os.path.join(ANTICIPOS_UPLOAD_FOLDER, filename)

        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas

        # Obtener datos de la empresa
        empresa = Empresa.query.first()
        
        c = canvas.Canvas(filepath, pagesize=A4)
        
        # Header con datos de empresa
        y_pos = 800
        if empresa:
            c.setFont('Helvetica-Bold', 12)
            c.drawString(50, y_pos, empresa.nombre or 'EMPRESA')
            y_pos -= 15
            
            c.setFont('Helvetica', 9)
            if empresa.ruc:
                c.drawString(50, y_pos, f'RUC: {empresa.ruc}')
                y_pos -= 12
            if empresa.direccion:
                c.drawString(50, y_pos, f'Dirección: {empresa.direccion}')
                y_pos -= 12
            if empresa.telefono:
                c.drawString(50, y_pos, f'Teléfono: {empresa.telefono}')
                y_pos -= 12
            if empresa.email:
                c.drawString(50, y_pos, f'Email: {empresa.email}')
                y_pos -= 12
        
        y_pos -= 10
        
        # Título
        c.setFont('Helvetica-Bold', 14)
        c.drawString(50, y_pos, 'RECIBO DE ANTICIPO')
        y_pos -= 20
        
        # Datos del empleado
        c.setFont('Helvetica', 10)
        empleado = anticipo.empleado
        c.drawString(50, y_pos, f'Empleado: {empleado.nombre_completo}')
        y_pos -= 15
        
        c.drawString(50, y_pos, f'CI: {empleado.ci}')
        y_pos -= 15
        
        c.drawString(50, y_pos, f'Monto: ₲ {float(anticipo.monto):,.0f}')
        y_pos -= 15
        
        c.drawString(50, y_pos, f'Fecha Aprobación: {anticipo.fecha_aprobacion.strftime("%d/%m/%Y %H:%M:%S")}')
        y_pos -= 15
        
        if anticipo.observaciones:
            c.drawString(50, y_pos, f'Observaciones: {anticipo.observaciones}')
            y_pos -= 15
        
        y_pos -= 20
        
        # Firmas
        c.drawString(50, y_pos, 'Firma Empleado: _______________________')
        y_pos -= 30
        
        c.drawString(50, y_pos, 'Firma RRHH: ___________________________')
        
        if empresa and empresa.representante_legal:
            y_pos -= 30
            c.drawString(50, y_pos, f'Firma Representante: _______________________')
            y_pos -= 15
            c.setFont('Helvetica', 8)
            c.drawString(50, y_pos, f'{empresa.representante_legal}')
        
        c.save()

        anticipo.justificativo_archivo = f"anticipos/{filename}"
        db.session.commit()

        registrar_operacion_crud(current_user, 'anticipos', 'UPDATE', 'anticipos', anticipo.id, {'aprobado': True})

        return jsonify({'message': 'anticipo aprobado', 'ruta_pdf': anticipo.justificativo_archivo}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@rrhh_bp.route('/anticipos/<int:anticipo_id>/reject', methods=['POST'])
@login_required
def rechazar_anticipo(anticipo_id):
    """Rechazar anticipo"""
    anticipo = Anticipo.query.get_or_404(anticipo_id)
    try:
        anticipo.rechazado = True
        anticipo.rechazado_por = current_user.id
        anticipo.fecha_rechazo = datetime.utcnow()
        db.session.commit()

        registrar_operacion_crud(current_user, 'anticipos', 'UPDATE', 'anticipos', anticipo.id, {'rechazado': True})

        return jsonify({'message': 'anticipo rechazado'}), 200
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error rechazando anticipo: {str(e)}")
        return jsonify({'error': str(e)}), 500


@rrhh_bp.route('/anticipos/<int:anticipo_id>/upload-justificativo', methods=['POST'])
@login_required
def upload_anticipo_justificativo(anticipo_id):
    """Subir justificativo para anticipo"""
    anticipo = Anticipo.query.get_or_404(anticipo_id)

    if 'file' in request.files:
        file = request.files['file']
    elif 'justificativo' in request.files:
        file = request.files['justificativo']
    else:
        return jsonify({'error': 'No se recibió ningún archivo (campo "file" o "justificativo" esperado)'}), 400

    if not file or file.filename == '':
        return jsonify({'error': 'No se seleccionó ningún archivo'}), 400

    filename_clean = secure_filename(file.filename)
    if not allowed_file(filename_clean):
        return jsonify({'error': 'Extensión de archivo no permitida'}), 400

    os.makedirs(ANTICIPOS_UPLOAD_FOLDER, exist_ok=True)
    filename = secure_filename(f"antic_{anticipo_id}_{int(datetime.utcnow().timestamp())}_{filename_clean}")
    filepath = os.path.join(ANTICIPOS_UPLOAD_FOLDER, filename)
    file.save(filepath)

    anticipo.justificativo_archivo = f"anticipos/{filename}"
    db.session.commit()

    registrar_operacion_crud(current_user, 'anticipos', 'UPDATE', 'anticipos', anticipo.id, {'justificativo': anticipo.justificativo_archivo})

    return jsonify({'message': 'uploaded', 'ruta': anticipo.justificativo_archivo}), 200


@rrhh_bp.route('/postulantes')
@login_required
@role_required(RoleEnum.RRHH)
def postulantes_lista():
    """Listado de postulantes con búsqueda y filtros"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    estado = request.args.get('estado', '', type=str)
    
    query = Postulante.query
    
    if search:
        query = query.filter(
            or_(
                Postulante.nombre.ilike(f'%{search}%'),
                Postulante.apellido.ilike(f'%{search}%'),
                Postulante.email.ilike(f'%{search}%'),
                Postulante.cargo_postulado.ilike(f'%{search}%')
            )
        )
    
    if estado:
        query = query.filter_by(estado=estado)
    
    postulantes = query.order_by(desc(Postulante.fecha_postulacion)).paginate(page=page, per_page=15)
    
    return render_template('rrhh/postulantes_lista.html', postulantes=postulantes, search=search, estado=estado)


@rrhh_bp.route('/postulantes/nuevo', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def postulante_nuevo():
    """Crear un nuevo postulante"""
    if request.method == 'POST':
        # Validar datos básicos
        nombre = request.form.get('nombre', '').strip()
        apellido = request.form.get('apellido', '').strip()
        email = request.form.get('email', '').strip()
        telefono = request.form.get('telefono', '').strip()
        fecha_nacimiento_str = request.form.get('fecha_nacimiento', '')
        nivel_academico = request.form.get('nivel_academico', '').strip()
        cargo_postulado = request.form.get('cargo_postulado', '').strip()
        experiencia_años = request.form.get('experiencia_años', 0, type=int)
        salario_esperado = request.form.get('salario_esperado', None, type=float)
        observaciones = request.form.get('observaciones', '').strip()
        
        if not nombre or not apellido or not email or not cargo_postulado:
            flash('Nombre, apellido, email y cargo postulado son obligatorios', 'danger')
            return redirect(url_for('rrhh.postulante_nuevo'))
        
        # Verificar email único
        if Postulante.query.filter_by(email=email).first():
            flash('Este email ya está registrado como postulante', 'danger')
            return redirect(url_for('rrhh.postulante_nuevo'))
        
        # Crear postulante
        try:
            fecha_nacimiento = None
            if fecha_nacimiento_str:
                fecha_nacimiento = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
            
            postulante = Postulante(
                nombre=nombre,
                apellido=apellido,
                email=email,
                telefono=telefono,
                fecha_nacimiento=fecha_nacimiento,
                nivel_academico=nivel_academico,
                cargo_postulado=cargo_postulado,
                experiencia_años=experiencia_años,
                salario_esperado=Decimal(str(salario_esperado)) if salario_esperado else None,
                estado='Nuevo',
                observaciones=observaciones,
                fecha_postulacion=date.today()
            )
            db.session.add(postulante)
            db.session.flush()  # Para obtener el ID antes de commit
            
            # Procesar archivos (CV, certificados, etc.)
            if 'documentos' in request.files:
                files = request.files.getlist('documentos')
                for file in files:
                    if file and file.filename != '':
                        if not allowed_file(file.filename):
                            flash(f'Archivo {file.filename} no permitido (solo PNG, JPG, PDF)', 'warning')
                            continue
                        
                        os.makedirs(POSTULANTES_UPLOAD_FOLDER, exist_ok=True)
                        filename_clean = secure_filename(file.filename)
                        filename = secure_filename(f"post_{postulante.id}_{int(datetime.utcnow().timestamp())}_{filename_clean}")
                        filepath = os.path.join(POSTULANTES_UPLOAD_FOLDER, filename)
                        file.save(filepath)
                        
                        # Determinar tipo de documento
                        ext = filename_clean.rsplit('.', 1)[1].lower()
                        tipo = 'CV' if filename_clean.lower().startswith('cv') else 'Certificado'
                        
                        doc = DocumentosCurriculum(
                            postulante_id=postulante.id,
                            tipo=tipo,
                            nombre_archivo=filename_clean,
                            ruta_archivo=os.path.relpath(filepath, os.path.dirname(os.path.dirname(__file__))),
                            tamaño_bytes=len(file.read()),
                            mime_type=file.content_type or 'application/octet-stream'
                        )
                        file.seek(0)  # Reset file pointer
                        tamaño_bytes = len(file.read())
                        doc.tamaño_bytes = tamaño_bytes
                        db.session.add(doc)
            
            db.session.commit()
            registrar_operacion_crud(current_user, 'postulantes', 'CREATE', 'postulantes', postulante.id, {'nombre': nombre, 'apellido': apellido, 'email': email})
            
            flash(f'Postulante {nombre} {apellido} creado exitosamente', 'success')
            return redirect(url_for('rrhh.postulante_detalle', postulante_id=postulante.id))
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear postulante: {str(e)}', 'danger')
            return redirect(url_for('rrhh.postulante_nuevo'))
    
    return render_template('rrhh/postulante_form.html')


@rrhh_bp.route('/postulantes/<int:postulante_id>')
@login_required
@role_required(RoleEnum.RRHH)
def postulante_detalle(postulante_id):
    """Detalle de un postulante"""
    postulante = Postulante.query.get_or_404(postulante_id)
    return render_template('rrhh/postulante_detalle.html', postulante=postulante)


@rrhh_bp.route('/postulantes/<int:postulante_id>/cambiar-estado', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def postulante_cambiar_estado(postulante_id):
    """Cambiar estado del postulante"""
    postulante = Postulante.query.get_or_404(postulante_id)
    nuevo_estado = request.form.get('estado', '').strip()
    
    estados_validos = ['Nuevo', 'En Evaluación', 'Contratado', 'Rechazado', 'En Espera']
    if nuevo_estado not in estados_validos:
        flash('Estado inválido', 'danger')
        return redirect(url_for('rrhh.postulante_detalle', postulante_id=postulante_id))
    
    postulante.estado = nuevo_estado
    postulante.fecha_actualizado = datetime.utcnow()
    db.session.commit()
    
    registrar_operacion_crud(current_user, 'postulantes', 'UPDATE', 'postulantes', postulante.id, {'estado': nuevo_estado})
    flash(f'Estado del postulante actualizado a: {nuevo_estado}', 'success')
    
    return redirect(url_for('rrhh.postulante_detalle', postulante_id=postulante_id))


@rrhh_bp.route('/postulantes/<int:postulante_id>/descargar-documento/<int:doc_id>')
@login_required
@role_required(RoleEnum.RRHH)
def descargar_documento_postulante(postulante_id, doc_id):
    """Descargar documento de curriculum"""
    postulante = Postulante.query.get_or_404(postulante_id)
    doc = DocumentosCurriculum.query.get_or_404(doc_id)
    
    if doc.postulante_id != postulante_id:
        flash('Acceso denegado', 'danger')
        return redirect(url_for('rrhh.postulantes_lista'))
    
    try:
        filepath = os.path.join(os.path.dirname(os.path.dirname(__file__)), doc.ruta_archivo)
        return send_file(filepath, as_attachment=True, download_name=doc.nombre_archivo)
    except Exception as e:
        flash(f'Error al descargar archivo: {str(e)}', 'danger')
        return redirect(url_for('rrhh.postulante_detalle', postulante_id=postulante_id))


# ----------------- CONTRATOS (GENERACIÓN con ReportLab y almacenamiento en DB) -----------------


def _generar_numero_contrato(empleado_id):
    """Genera un número de contrato único: CONT-YYYY-EMPID-TIMESTAMP"""
    ts = int(datetime.utcnow().timestamp())
    year = datetime.utcnow().year
    return f'CONT-{year}-{empleado_id}-{ts}'


def generar_pdf_contrato(datos: dict) -> bytes:
    """Genera un PDF de contrato usando ReportLab y devuelve bytes.

    datos: dict con claves: contratante, contratado, ci, direccion_contratante, direccion_contratado,
    objeto, fecha_inicio (date), fecha_fin (date), monto, ciudad, dia, mes, ano
    """
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Margenes
    margin_left = 50
    y = height - 70

    c.setFont('Helvetica-Bold', 14)
    c.drawString(margin_left, y, 'CONTRATO DE PRESTACIÓN DE SERVICIOS')
    y -= 30

    c.setFont('Helvetica', 11)

    # Plantilla (lineas, usando el texto provisto por el usuario)
    plantilla = (
        "Entre {contratante}, con domicilio en {direccion_contratante}, en adelante “EL CONTRATANTE”,"
        " y {contratado}, con cédula de identidad Nº {ci}, con domicilio en {direccion_contratado},"
        " en adelante “EL CONTRATADO”, se celebra el presente contrato conforme a las siguientes cláusulas:\n\n"
        "PRIMERA – Objeto\n\n"
        "EL CONTRATADO se compromete a prestar servicios de {objeto} a favor de EL CONTRATANTE.\n\n"
        "SEGUNDA – Plazo\n\n"
        "El presente contrato tendrá una duración de {meses} meses, iniciando el {fecha_inicio} y finalizando el {fecha_fin}, pudiendo renovarse por acuerdo entre las partes.\n\n"
        "TERCERA – Honorarios\n\n"
        "EL CONTRATANTE abonará a EL CONTRATADO la suma de {monto}, pagadera en {periodicidad}.\n\n"
        "CUARTA – Obligaciones\n\n"
        "EL CONTRATADO se obliga a cumplir con las tareas asignadas con responsabilidad, confidencialidad y dentro de los plazos establecidos.\n"
        "EL CONTRATANTE se compromete a proporcionar los medios y la información necesarios para la correcta ejecución del servicio.\n\n"
        "QUINTA – Terminación\n\n"
        "Cualquiera de las partes podrá rescindir el presente contrato con un preaviso de {preaviso} días, sin derecho a indemnización, salvo los honorarios devengados hasta la fecha.\n\n"
        "SEXTA – Jurisdicción\n\n"
        "Para cualquier controversia derivada del presente contrato, las partes se someten a la jurisdicción de los tribunales de {ciudad}.\n\n"
        "En prueba de conformidad, se firman dos (2) ejemplares de un mismo tenor y efecto, en {ciudad}, a los {dia} días del mes de {mes} de {ano}.\n\n"
        "EL CONTRATANTE\n\nFirma: ______________________\nNombre: {contratante}\n\n"
        "EL CONTRATADO\n\nFirma: ______________________\nNombre: {contratado}\n"
    )

    texto = plantilla.format(**datos)

    # Dibujar texto con wrapping simple
    textobj = c.beginText()
    textobj.setTextOrigin(margin_left, y)
    textobj.setLeading(14)
    for line in texto.split('\n'):
        # dividir si la linea es muy larga
        if len(line) > 120:
            # simple wrap
            while len(line) > 120:
                textobj.textLine(line[:120])
                line = line[120:]
            if line:
                textobj.textLine(line)
        else:
            textobj.textLine(line)

    c.drawText(textobj)
    c.showPage()
    c.save()

    buffer.seek(0)
    return buffer.read()


@rrhh_bp.route('/empleados/<int:empleado_id>/contratos/nuevo', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def contrato_nuevo(empleado_id):
    empleado = Empleado.query.get_or_404(empleado_id)
    if request.method == 'POST':
        tipo = request.form.get('tipo_contrato', 'Temporal')
        fecha_inicio_str = request.form.get('fecha_inicio', '')
        fecha_fin_str = request.form.get('fecha_fin', '')
        meses = request.form.get('meses', '0')
        monto = request.form.get('monto', '0')
        periodicidad = request.form.get('periodicidad', 'mensual')
        preaviso = request.form.get('preaviso', '30')

        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date() if fecha_inicio_str else None
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() if fecha_fin_str else None

            numero = _generar_numero_contrato(empleado_id)

            datos_pdf = {
                'contratante': request.form.get('contratante_nombre', 'EMPRESA'),
                'direccion_contratante': request.form.get('direccion_contratante', ''),
                'contratado': f"{empleado.nombre} {empleado.apellido}",
                'direccion_contratado': request.form.get('direccion_contratado', empleado.direccion if hasattr(empleado, 'direccion') else ''),
                'ci': request.form.get('ci', getattr(empleado, 'ci', '')),
                'objeto': request.form.get('objeto', 'servicios'),
                'meses': meses,
                'fecha_inicio': fecha_inicio.strftime('%d/%m/%Y') if fecha_inicio else '',
                'fecha_fin': fecha_fin.strftime('%d/%m/%Y') if fecha_fin else '',
                'monto': monto,
                'periodicidad': periodicidad,
                'preaviso': preaviso,
                'ciudad': request.form.get('ciudad', 'Ciudad'),
                'dia': fecha_inicio.day if fecha_inicio else datetime.utcnow().day,
                'mes': fecha_inicio.strftime('%B') if fecha_inicio else datetime.utcnow().strftime('%B'),
                'ano': fecha_inicio.year if fecha_inicio else datetime.utcnow().year
            }

            pdf_bytes = generar_pdf_contrato(datos_pdf)

            contrato = Contrato(
                empleado_id=empleado_id,
                numero_contrato=numero,
                tipo_contrato=tipo,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                contenido=pdf_bytes,
                variables=json.dumps(datos_pdf, default=str)
            )
            db.session.add(contrato)
            db.session.commit()

            registrar_operacion_crud(current_user, 'contratos', 'CREATE', 'contratos', contrato.id, {'numero': numero})
            flash('Contrato generado y guardado correctamente', 'success')
            return redirect(url_for('rrhh.perfil_empleado', empleado_id=empleado_id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear contrato: {str(e)}', 'danger')
            return redirect(url_for('rrhh.contrato_nuevo', empleado_id=empleado_id))

    # GET
    return render_template('rrhh/contrato_form.html', empleado=empleado)


@rrhh_bp.route('/contratos/<int:contrato_id>/descargar')
@login_required
@role_required(RoleEnum.RRHH)
def contrato_descargar(contrato_id):
    contrato = Contrato.query.get_or_404(contrato_id)
    if not contrato.contenido:
        flash('No hay contenido PDF para este contrato', 'danger')
        return redirect(url_for('rrhh.perfil_empleado', empleado_id=contrato.empleado_id))
    buf = BytesIO(contrato.contenido)
    filename = f"{contrato.numero_contrato}.pdf"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')


# ==================== BONIFICACIÓN FAMILIAR ====================

def obtener_salario_minimo_vigente(fecha=None):
    """
    Obtiene el salario mínimo vigente para una fecha específica.
    Si no se proporciona fecha, usa la fecha actual.
    """
    if fecha is None:
        fecha = date.today()
    
    salario = SalarioMinimo.query.filter(
        SalarioMinimo.vigencia_desde <= fecha,
        or_(SalarioMinimo.vigencia_hasta.is_(None), SalarioMinimo.vigencia_hasta >= fecha)
    ).order_by(desc(SalarioMinimo.vigencia_desde)).first()
    
    if not salario:
        # Si no hay salario mínimo registrado, retornar el último conocido
        salario = SalarioMinimo.query.order_by(desc(SalarioMinimo.vigencia_desde)).first()
    
    return salario.monto if salario else Decimal('2798309')  # Fallback a salario mínimo 2025

def contar_hijos_activos(empleado_id, fecha=None):
    """
    Cuenta la cantidad de hijos activos de un empleado para bonificación familiar.
    """
    if fecha is None:
        fecha = date.today()
    
    return BonificacionFamiliar.query.filter(
        BonificacionFamiliar.empleado_id == empleado_id,
        BonificacionFamiliar.activo == True,
        or_(
            BonificacionFamiliar.fecha_baja.is_(None),
            BonificacionFamiliar.fecha_baja >= fecha
        )
    ).count()

def calcular_bonificacion_familiar(empleado_id, fecha=None):
    """
    Calcula bonificación familiar según ley paraguaya.
    Fórmula: (Salario Mínimo × 5%) × Cantidad de hijos activos
    """
    if fecha is None:
        fecha = date.today()
    
    # Obtener salario mínimo vigente
    salario_minimo = obtener_salario_minimo_vigente(fecha)
    
    # Contar hijos activos
    hijos_activos = contar_hijos_activos(empleado_id, fecha)
    
    if hijos_activos == 0:
        return Decimal('0')
    
    # Calcular bonificación
    bonificacion_por_hijo = (salario_minimo * Decimal('0.05')).quantize(Decimal('0.01'))
    bonificacion_total = (bonificacion_por_hijo * hijos_activos).quantize(Decimal('0.01'))
    
    return bonificacion_total

# ==================== RUTAS: GESTIÓN DE SALARIOS MÍNIMOS ====================

@rrhh_bp.route('/salarios_minimos', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def listar_salarios_minimos():
    """Lista histórico de salarios mínimos"""
    registrar_bitacora(current_user, 'salarios_minimos', 'VIEW', 'salarios_minimos')
    
    salarios = SalarioMinimo.query.order_by(desc(SalarioMinimo.vigencia_desde)).all()
    salario_vigente = obtener_salario_minimo_vigente()
    
    return render_template('rrhh/salarios_minimos.html', 
                          salarios=salarios,
                          salario_vigente=salario_vigente)

@rrhh_bp.route('/salarios_minimos/crear', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def crear_salario_minimo():
    """Crear nuevo salario mínimo"""
    if request.method == 'POST':
        try:
            año = int(request.form.get('año'))
            monto = Decimal(request.form.get('monto'))
            vigencia_desde = datetime.strptime(request.form.get('vigencia_desde'), '%Y-%m-%d').date()
            vigencia_hasta_str = request.form.get('vigencia_hasta')
            vigencia_hasta = datetime.strptime(vigencia_hasta_str, '%Y-%m-%d').date() if vigencia_hasta_str else None
            
            # Cerrar salario mínimo anterior si existe
            salario_anterior = SalarioMinimo.query.filter(
                SalarioMinimo.vigencia_hasta.is_(None)
            ).first()
            
            if salario_anterior and vigencia_desde > salario_anterior.vigencia_desde:
                salario_anterior.vigencia_hasta = vigencia_desde - timedelta(days=1)
            
            salario = SalarioMinimo(
                año=año,
                monto=monto,
                vigencia_desde=vigencia_desde,
                vigencia_hasta=vigencia_hasta,
                usuario_creador_id=current_user.id
            )
            
            db.session.add(salario)
            db.session.commit()
            
            registrar_operacion_crud(current_user, 'salarios_minimos', 'CREATE', 'salarios_minimos',
                                    salario.id, {'año': año, 'monto': str(monto)})
            
            flash(f'Salario mínimo {año} registrado: {monto} Gs.', 'success')
            return redirect(url_for('rrhh.listar_salarios_minimos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear salario mínimo: {str(e)}', 'danger')
    
    return render_template('rrhh/crear_salario_minimo.html')

# ==================== RUTAS: GESTIÓN DE HIJOS (BONIFICACIÓN FAMILIAR) ====================

@rrhh_bp.route('/empleado/<int:empleado_id>/hijos', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def listar_hijos(empleado_id):
    """Lista hijos de un empleado para bonificación familiar"""
    empleado = Empleado.query.get_or_404(empleado_id)
    registrar_bitacora(current_user, 'bonificacion_familiar', 'VIEW', 'bonificaciones_familiares')
    
    hijos = BonificacionFamiliar.query.filter_by(empleado_id=empleado_id).order_by(
        desc(BonificacionFamiliar.activo),
        BonificacionFamiliar.hijo_fecha_nacimiento
    ).all()
    
    # Calcular bonificación actual
    bonificacion_actual = calcular_bonificacion_familiar(empleado_id)
    salario_minimo = obtener_salario_minimo_vigente()
    hijos_activos_count = contar_hijos_activos(empleado_id)
    
    return render_template('rrhh/hijos_listado.html',
                          empleado=empleado,
                          hijos=hijos,
                          bonificacion_actual=bonificacion_actual,
                          salario_minimo=salario_minimo,
                          hijos_activos=hijos_activos_count)

@rrhh_bp.route('/empleado/<int:empleado_id>/hijos/agregar', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def agregar_hijo(empleado_id):
    """Agregar hijo para bonificación familiar"""
    empleado = Empleado.query.get_or_404(empleado_id)
    
    if request.method == 'POST':
        try:
            hijo_nombre = request.form.get('hijo_nombre')
            hijo_apellido = request.form.get('hijo_apellido')
            hijo_ci = request.form.get('hijo_ci')
            hijo_fecha_nacimiento = datetime.strptime(request.form.get('hijo_fecha_nacimiento'), '%Y-%m-%d').date()
            sexo = request.form.get('sexo')
            tipo_str = request.form.get('tipo')
            observaciones = request.form.get('observaciones')
            
            # Convertir tipo string a enum
            tipo = TipoHijoEnum[tipo_str.replace(' ', '_').replace('-', '_').upper()]
            
            # Manejar archivos subidos
            certificado_nacimiento = None
            certificado_estudio = None
            certificado_discapacidad = None
            
            upload_folder = os.path.join(current_app.static_folder, 'uploads', 'bonificaciones')
            os.makedirs(upload_folder, exist_ok=True)
            
            if 'certificado_nacimiento' in request.files:
                file = request.files['certificado_nacimiento']
                if file and file.filename:
                    filename = secure_filename(f"{empleado.codigo}_{hijo_nombre}_{hijo_apellido}_nacimiento_{file.filename}")
                    filepath = os.path.join(upload_folder, filename)
                    file.save(filepath)
                    certificado_nacimiento = f'uploads/bonificaciones/{filename}'
            
            if 'certificado_estudio' in request.files:
                file = request.files['certificado_estudio']
                if file and file.filename:
                    filename = secure_filename(f"{empleado.codigo}_{hijo_nombre}_{hijo_apellido}_estudio_{file.filename}")
                    filepath = os.path.join(upload_folder, filename)
                    file.save(filepath)
                    certificado_estudio = f'uploads/bonificaciones/{filename}'
            
            if 'certificado_discapacidad' in request.files:
                file = request.files['certificado_discapacidad']
                if file and file.filename:
                    filename = secure_filename(f"{empleado.codigo}_{hijo_nombre}_{hijo_apellido}_discapacidad_{file.filename}")
                    filepath = os.path.join(upload_folder, filename)
                    file.save(filepath)
                    certificado_discapacidad = f'uploads/bonificaciones/{filename}'
            
            hijo = BonificacionFamiliar(
                empleado_id=empleado_id,
                hijo_nombre=hijo_nombre,
                hijo_apellido=hijo_apellido,
                hijo_ci=hijo_ci,
                hijo_fecha_nacimiento=hijo_fecha_nacimiento,
                sexo=sexo,
                tipo=tipo,
                certificado_nacimiento=certificado_nacimiento,
                certificado_estudio=certificado_estudio,
                certificado_discapacidad=certificado_discapacidad,
                observaciones=observaciones
            )
            
            db.session.add(hijo)
            db.session.commit()
            
            registrar_operacion_crud(current_user, 'bonificacion_familiar', 'CREATE', 'bonificaciones_familiares',
                                    hijo.id, {'empleado': empleado.nombre_completo, 'hijo': hijo.hijo_nombre_completo})
            
            flash(f'Hijo {hijo.hijo_nombre_completo} agregado correctamente', 'success')
            return redirect(url_for('rrhh.listar_hijos', empleado_id=empleado_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al agregar hijo: {str(e)}', 'danger')
    
    return render_template('rrhh/agregar_hijo.html', empleado=empleado, tipos_hijo=TipoHijoEnum)

@rrhh_bp.route('/empleado/<int:empleado_id>/hijos/<int:hijo_id>/editar', methods=['GET', 'POST'])
@login_required
@role_required(RoleEnum.RRHH)
def editar_hijo(empleado_id, hijo_id):
    """Editar información de hijo"""
    empleado = Empleado.query.get_or_404(empleado_id)
    hijo = BonificacionFamiliar.query.get_or_404(hijo_id)
    
    if hijo.empleado_id != empleado_id:
        flash('El hijo no pertenece a este empleado', 'danger')
        return redirect(url_for('rrhh.listar_hijos', empleado_id=empleado_id))
    
    if request.method == 'POST':
        try:
            hijo.hijo_nombre = request.form.get('hijo_nombre')
            hijo.hijo_apellido = request.form.get('hijo_apellido')
            hijo.hijo_ci = request.form.get('hijo_ci')
            hijo.hijo_fecha_nacimiento = datetime.strptime(request.form.get('hijo_fecha_nacimiento'), '%Y-%m-%d').date()
            hijo.sexo = request.form.get('sexo')
            tipo_str = request.form.get('tipo')
            hijo.tipo = TipoHijoEnum[tipo_str.replace(' ', '_').replace('-', '_').upper()]
            hijo.observaciones = request.form.get('observaciones')
            
            # Actualizar archivos si se subieron nuevos
            upload_folder = os.path.join(current_app.static_folder, 'uploads', 'bonificaciones')
            os.makedirs(upload_folder, exist_ok=True)
            
            if 'certificado_nacimiento' in request.files:
                file = request.files['certificado_nacimiento']
                if file and file.filename:
                    filename = secure_filename(f"{empleado.codigo}_{hijo.hijo_nombre}_{hijo.hijo_apellido}_nacimiento_{file.filename}")
                    filepath = os.path.join(upload_folder, filename)
                    file.save(filepath)
                    hijo.certificado_nacimiento = f'uploads/bonificaciones/{filename}'
            
            if 'certificado_estudio' in request.files:
                file = request.files['certificado_estudio']
                if file and file.filename:
                    filename = secure_filename(f"{empleado.codigo}_{hijo.hijo_nombre}_{hijo.hijo_apellido}_estudio_{file.filename}")
                    filepath = os.path.join(upload_folder, filename)
                    file.save(filepath)
                    hijo.certificado_estudio = f'uploads/bonificaciones/{filename}'
            
            if 'certificado_discapacidad' in request.files:
                file = request.files['certificado_discapacidad']
                if file and file.filename:
                    filename = secure_filename(f"{empleado.codigo}_{hijo.hijo_nombre}_{hijo_apellido}_discapacidad_{file.filename}")
                    filepath = os.path.join(upload_folder, filename)
                    file.save(filepath)
                    hijo.certificado_discapacidad = f'uploads/bonificaciones/{filename}'
            
            db.session.commit()
            
            registrar_operacion_crud(current_user, 'bonificacion_familiar', 'UPDATE', 'bonificaciones_familiares',
                                    hijo.id, {'hijo': hijo.hijo_nombre_completo})
            
            flash(f'Información de {hijo.hijo_nombre_completo} actualizada', 'success')
            return redirect(url_for('rrhh.listar_hijos', empleado_id=empleado_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al editar hijo: {str(e)}', 'danger')
    
    return render_template('rrhh/editar_hijo.html', empleado=empleado, hijo=hijo, tipos_hijo=TipoHijoEnum)

@rrhh_bp.route('/empleado/<int:empleado_id>/hijos/<int:hijo_id>/dar_baja', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def dar_baja_hijo(empleado_id, hijo_id):
    """Dar de baja a un hijo (cumplió 18, terminó estudios, etc.)"""
    hijo = BonificacionFamiliar.query.get_or_404(hijo_id)
    
    if hijo.empleado_id != empleado_id:
        flash('El hijo no pertenece a este empleado', 'danger')
        return redirect(url_for('rrhh.listar_hijos', empleado_id=empleado_id))
    
    try:
        motivo = request.form.get('motivo', 'Baja manual')
        hijo.activo = False
        hijo.fecha_baja = date.today()
        hijo.motivo_baja = motivo
        
        db.session.commit()
        
        registrar_operacion_crud(current_user, 'bonificacion_familiar', 'UPDATE', 'bonificaciones_familiares',
                                hijo.id, {'accion': 'Dar de baja', 'motivo': motivo})
        
        flash(f'Hijo {hijo.hijo_nombre_completo} dado de baja', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al dar de baja: {str(e)}', 'danger')
    
    return redirect(url_for('rrhh.listar_hijos', empleado_id=empleado_id))

@rrhh_bp.route('/empleado/<int:empleado_id>/hijos/<int:hijo_id>/reactivar', methods=['POST'])
@login_required
@role_required(RoleEnum.RRHH)
def reactivar_hijo(empleado_id, hijo_id):
    """Reactivar un hijo dado de baja"""
    hijo = BonificacionFamiliar.query.get_or_404(hijo_id)
    
    if hijo.empleado_id != empleado_id:
        flash('El hijo no pertenece a este empleado', 'danger')
        return redirect(url_for('rrhh.listar_hijos', empleado_id=empleado_id))
    
    try:
        hijo.activo = True
        hijo.fecha_baja = None
        hijo.motivo_baja = None
        
        db.session.commit()
        
        registrar_operacion_crud(current_user, 'bonificacion_familiar', 'UPDATE', 'bonificaciones_familiares',
                                hijo.id, {'accion': 'Reactivar'})
        
        flash(f'Hijo {hijo.hijo_nombre_completo} reactivado', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al reactivar: {str(e)}', 'danger')
    
    return redirect(url_for('rrhh.listar_hijos', empleado_id=empleado_id))

@rrhh_bp.route('/bonificaciones_familiares/reporte', methods=['GET'])
@login_required
@role_required(RoleEnum.RRHH)
def reporte_bonificaciones():
    """Reporte consolidado de bonificaciones familiares"""
    registrar_bitacora(current_user, 'bonificacion_familiar', 'VIEW', 'bonificaciones_familiares')
    
    # Obtener todos los empleados activos con hijos
    empleados_con_hijos = db.session.query(
        Empleado,
        func.count(BonificacionFamiliar.id).label('cantidad_hijos')
    ).join(
        BonificacionFamiliar, Empleado.id == BonificacionFamiliar.empleado_id
    ).filter(
        Empleado.estado == EstadoEmpleadoEnum.ACTIVO,
        BonificacionFamiliar.activo == True
    ).group_by(Empleado.id).all()
    
    # Calcular bonificaciones
    reporte = []
    total_bonificaciones = Decimal('0')
    total_hijos = 0
    
    for empleado, cant_hijos in empleados_con_hijos:
        bonificacion = calcular_bonificacion_familiar(empleado.id)
        reporte.append({
            'empleado': empleado,
            'cantidad_hijos': cant_hijos,
            'bonificacion': bonificacion
        })
        total_bonificaciones += bonificacion
        total_hijos += cant_hijos
    
    salario_minimo = obtener_salario_minimo_vigente()
    bonificacion_por_hijo = (salario_minimo * Decimal('0.05')).quantize(Decimal('0.01'))
    
    return render_template('rrhh/reporte_bonificaciones.html',
                          reporte=reporte,
                          total_bonificaciones=total_bonificaciones,
                          total_hijos=total_hijos,
                          salario_minimo=salario_minimo,
                          bonificacion_por_hijo=bonificacion_por_hijo)


